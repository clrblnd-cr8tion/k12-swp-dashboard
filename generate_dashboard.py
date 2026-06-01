#!/usr/bin/env python3
"""
Generate an interactive HTML dashboard from the K12 SWP Excel spreadsheet.

Reads Round 5–8 tabs and produces a self-contained dashboard.html with
5 switchable views: Command Center, Portfolio Overview, Status Matrix,
Alert Dashboard, and Grant Cards.

Usage: python3 generate_dashboard.py
Output: dashboard.html
"""
import base64
import json
import re
import os
from datetime import date, datetime
from openpyxl import load_workbook

from utils import SPREADSHEET, SCRAPED_DATA, is_data_row

# ---------------------------------------------------------------------------
# Round metadata (quarterly deadlines hardcoded per CLAUDE.md)
# ---------------------------------------------------------------------------
ROUND_META = {
    'R5': {
        'sheet': 'Round 5', 'label': 'Round 5', 'grantFY': '2022-23',
        'yearCode': '2023004', 'finalReportDue': '2025-09-30',
        'quarterlyDueDates': ['2023-08-31', '2024-02-28', '2024-08-31', '2025-02-28', '2025-08-31'],
        'grantFYStart': '2022-07-01',
    },
    'R6': {
        'sheet': 'Round 6', 'label': 'Round 6', 'grantFY': '2023-24',
        'yearCode': '2024004', 'finalReportDue': '2026-09-30',
        'quarterlyDueDates': ['2024-08-31', '2025-02-28', '2025-08-31', '2026-02-28', '2026-08-31'],
        'grantFYStart': '2023-07-01',
    },
    'R7': {
        'sheet': 'Round 7', 'label': 'Round 7', 'grantFY': '2024-25',
        'yearCode': '2025004', 'finalReportDue': '2027-09-30',
        'quarterlyDueDates': ['2025-08-31', '2026-02-28', '2026-08-31', '2027-02-28', '2027-08-31'],
        'grantFYStart': '2024-07-01',
    },
    'R8': {
        'sheet': 'Round 8', 'label': 'Round 8', 'grantFY': '2025-26',
        'yearCode': '2026004', 'finalReportDue': '2028-09-30',
        'quarterlyDueDates': ['2026-08-31', '2027-02-28', '2027-08-31', '2028-02-28', '2028-08-31'],
        'grantFYStart': '2025-07-01',
    },
}


def norm(v):
    """Normalize a cell value: convert to string, replace newlines with space, strip."""
    if v is None:
        return ''
    return str(v).replace('\n', ' ').strip()


def load_scraped():
    """Load scraped_data.json indexed by planId. Returns {} if not found."""
    if not os.path.exists(SCRAPED_DATA):
        return {}
    with open(SCRAPED_DATA, 'r') as f:
        raw = json.load(f)
    index = {}
    for rd in raw.values():
        grants = rd if isinstance(rd, list) else rd.get('grants', [])
        for g in grants:
            index[str(g['planId'])] = g
    return index


def extract_round(ws, rk, meta, today, scraped):
    """Extract grants from one round worksheet. Returns a round data dict."""
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))

    # Build header→col_index map (0-based)
    col_map = {}
    quarterly_col_indices = []  # 0-based indices of E–I quarterly columns
    quarterly_labels_raw = []
    final_col = None

    for i, cell in enumerate(header_row):
        h = norm(cell)
        if not h:
            continue
        col_map[h] = i
        if re.search(r'FY\d{2}-\d{2}\s+Q[24]', h):
            quarterly_col_indices.append(i)
            quarterly_labels_raw.append(h)
        if h.startswith('Final Report'):
            final_col = i

    quarterly_col_indices.sort()

    # Locate key columns by exact or partial match
    def exact(name):
        return col_map.get(name)

    def partial(*terms):
        for h, idx in col_map.items():
            hl = h.lower()
            if all(t.lower() in hl for t in terms):
                return idx
        return None

    budget_col = exact('Grant Amount')
    # Match "Total Reported Expenditures" but NOT "Approved"
    exp_col = None
    for h, idx in col_map.items():
        if 'total reported expenditures' in h.lower() and 'approved' not in h.lower():
            exp_col = idx
            break
    inst_count_col = partial('Reporting')   # "# Reporting Insitutions" (typo is intentional)
    plan_status_col = exact('Plan Status')
    lead_lea_col = exact('Lead LEA')
    waiting_col = exact('Report Waiting Approval')

    due_dates = [date.fromisoformat(d) for d in meta['quarterlyDueDates']]
    final_due = date.fromisoformat(meta['finalReportDue'])

    grants = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        pid = row[1]  # Column B: Proposal ID
        if not is_data_row(pid):
            continue

        pid_str = str(pid)
        institution = norm(row[0])   # Column A
        grant_name = norm(row[3])    # Column D

        raw_count = row[inst_count_col] if inst_count_col is not None else 1
        raw_budget = row[budget_col] if budget_col is not None else 0
        raw_spent = row[exp_col] if exp_col is not None else 0

        budget = float(raw_budget) if raw_budget else 0.0
        spent = float(raw_spent) if raw_spent else 0.0
        pct_spent = (spent / budget) if budget > 0 else 0.0

        plan_status = norm(row[plan_status_col]) if plan_status_col is not None else ''
        lead_lea = norm(row[lead_lea_col]) if lead_lea_col is not None else ''
        waiting = norm(row[waiting_col]) if waiting_col is not None else ''

        # Quarterly booleans (True / None in xlsx)
        quarterly = []
        for ci in quarterly_col_indices:
            val = row[ci] if ci < len(row) else None
            quarterly.append(bool(val) if val is not None else False)

        # Final report
        final_val = row[final_col] if final_col is not None and final_col < len(row) else None
        final_report = bool(final_val) if final_val is not None else False

        # At-risk: any past-due quarter not submitted, or final report past due
        at_risk = any(not q and today > d for q, d in zip(quarterly, due_dates))
        if not final_report and today > final_due:
            at_risk = True

        # Supplement from scraped_data.json for extra fields
        sg = scraped.get(pid_str, {})
        if not plan_status:
            plan_status = sg.get('planStatus', '')
        if not lead_lea:
            lead_lea = sg.get('leadLEA', '')
        dashboard_approval = sg.get('dashboardApproval', '') or waiting
        approval_status = sg.get('approvalStatus', '')

        grants.append({
            'planId': pid_str,
            'institution': institution,
            'grantName': grant_name,
            'instCount': int(raw_count) if raw_count else 1,
            'budget': round(budget, 2),
            'spent': round(spent, 2),
            'pctSpent': round(pct_spent, 4),
            'planStatus': plan_status,
            'leadLEA': lead_lea,
            'dashboardApproval': dashboard_approval,
            'approvalStatus': approval_status,
            'certifiedDate': sg.get('certifiedDate', ''),
            'submittedDate': sg.get('submittedDate', ''),
            'quarterly': quarterly,
            'finalReport': final_report,
            'atRisk': at_risk,
        })

    total_budget = sum(g['budget'] for g in grants)
    total_spent = sum(g['spent'] for g in grants)
    n = len(grants)

    return {
        'label': meta['label'],
        'grantFY': meta['grantFY'],
        'yearCode': meta['yearCode'],
        'finalReportDue': meta['finalReportDue'],
        'grantFYStart': meta['grantFYStart'],
        'quarterlyLabels': quarterly_labels_raw,
        'quarterlyDueDates': meta['quarterlyDueDates'],
        'grants': grants,
        'totals': {
            'budget': round(total_budget, 2),
            'spent': round(total_spent, 2),
            'pctSpent': round(total_spent / total_budget, 4) if total_budget > 0 else 0,
            'grantCount': n,
            'atRiskCount': sum(1 for g in grants if g['atRisk']),
            'quarterlyCounts': [sum(1 for g in grants if g['quarterly'][i])
                                for i in range(len(quarterly_col_indices))],
            'finalCount': sum(1 for g in grants if g['finalReport']),
        },
    }


# ---------------------------------------------------------------------------
# HTML Template
# ---------------------------------------------------------------------------
HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>K12 SWP Grant Dashboard — Central Mother Lode</title>
<script>
(function(){
  if(sessionStorage.getItem('swp_auth')==='1') return;
  document.addEventListener('DOMContentLoaded',function(){
    var gate=document.getElementById('pw-gate');
    var inp=document.getElementById('pw-input');
    var err=document.getElementById('pw-err');
    document.getElementById('pw-form').addEventListener('submit',function(e){
      e.preventDefault();
      if(inp.value==='strongworkforce2026'){
        sessionStorage.setItem('swp_auth','1');
        gate.style.display='none';
      } else {
        err.style.display='block';
        inp.value='';
        inp.focus();
      }
    });
  });
})();
</script>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@400;500;600;700;800&display=swap" rel="stylesheet">
<style>
:root {
  --navy:       #1B3A8F;
  --navy-light: #2952C8;
  --blue:       #3B82F6;
  --blue-bg:    #EFF6FF;
  --blue-mid:   #BFDBFE;
  --green:      #059669;
  --green-bg:   #D1FAE5;
  --amber:      #D97706;
  --amber-bg:   #FEF3C7;
  --red:        #DC2626;
  --red-bg:     #FEE2E2;
  --text:       #0F172A;
  --muted:      #64748B;
  --border:     #E2E8F0;
  --bg:         #F1F5FB;
  --card:       #FFFFFF;
  --r:          10px;
  --font: 'Plus Jakarta Sans', system-ui, -apple-system, sans-serif;
}
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:var(--font);background:var(--bg);color:var(--text);font-size:14px;line-height:1.5}

/* ─── HEADER ─────────────────────────────────────────────── */
.hdr{background:var(--navy);color:#fff;position:sticky;top:0;z-index:200;box-shadow:0 4px 24px rgba(0,0,0,.18)}
.hdr-top{display:flex;align-items:center;justify-content:space-between;padding:12px 24px 10px;border-bottom:1px solid rgba(255,255,255,.1)}
.hdr-title{font-size:16px;font-weight:800;letter-spacing:-.3px;display:flex;align-items:center;gap:10px}
.rgn-badge{background:rgba(255,255,255,.15);padding:2px 10px;border-radius:100px;font-size:11px;font-weight:700;letter-spacing:.5px;text-transform:uppercase}
.hdr-meta{font-size:12px;color:rgba(255,255,255,.55)}

/* tabs */
.tabs{display:flex;padding:0 16px;border-bottom:1px solid rgba(255,255,255,.1);overflow-x:auto}
.tab{background:none;border:none;color:rgba(255,255,255,.55);font-family:var(--font);font-size:13px;font-weight:600;padding:10px 15px;cursor:pointer;position:relative;white-space:nowrap;transition:color .15s}
.tab:hover{color:rgba(255,255,255,.85)}
.tab.on{color:#fff}
.tab.on::after{content:'';position:absolute;bottom:0;left:0;right:0;height:3px;background:#60A5FA;border-radius:3px 3px 0 0}

/* filter bar */
.fbar{display:flex;align-items:center;gap:10px;padding:9px 24px;background:#fff;border-bottom:1px solid var(--border);flex-wrap:wrap}
.pills{display:flex;gap:5px}
.pill{background:var(--bg);border:1.5px solid var(--border);color:var(--muted);font-family:var(--font);font-size:12px;font-weight:700;padding:4px 13px;border-radius:100px;cursor:pointer;transition:all .15s}
.pill:hover{border-color:var(--navy);color:var(--navy)}
.pill.on{background:var(--navy);border-color:var(--navy);color:#fff}
.srch{flex:1;max-width:280px;padding:6px 13px;border:1.5px solid var(--border);border-radius:8px;font-family:var(--font);font-size:13px;color:var(--text);outline:none;transition:border-color .15s}
.srch:focus{border-color:var(--blue);box-shadow:0 0 0 3px rgba(59,130,246,.1)}
.rcount{margin-left:auto;font-size:12px;color:var(--muted);font-weight:500;white-space:nowrap}

/* ─── MAIN ────────────────────────────────────────────────── */
.main{padding:20px 24px;max-width:1700px;margin:0 auto}
.view{display:none;animation:fadeUp .2s ease}
.view.on{display:block}
@keyframes fadeUp{from{opacity:0;transform:translateY(6px)}to{opacity:1;transform:translateY(0)}}

/* ─── KPI CARDS ──────────────────────────────────────────── */
.kpi-row{display:grid;grid-template-columns:repeat(4,1fr);gap:16px;margin-bottom:20px}
.kpi{background:var(--card);border-radius:var(--r);padding:18px 20px;border:1px solid var(--border);box-shadow:0 1px 4px rgba(0,0,0,.04)}
.kpi-lbl{font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.7px;color:var(--muted);margin-bottom:6px}
.kpi-val{font-size:28px;font-weight:800;letter-spacing:-.5px;line-height:1;margin-bottom:3px}
.kpi-sub{font-size:12px;color:var(--muted)}
.kpi.red .kpi-val{color:var(--red)}
.kpi.grn .kpi-val{color:var(--green)}
.kpi.nvy .kpi-val{color:var(--navy)}

/* ─── TABLE ──────────────────────────────────────────────── */
.tbl-wrap{background:var(--card);border-radius:var(--r);border:1px solid var(--border);overflow:auto;box-shadow:0 1px 4px rgba(0,0,0,.04)}
table{width:100%;border-collapse:collapse;font-size:13px}
thead th{background:#F8FAFC;padding:10px 13px;text-align:left;font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:.5px;color:var(--muted);border-bottom:1px solid var(--border);white-space:nowrap;cursor:pointer;user-select:none;position:sticky;top:0}
thead th:hover{color:var(--navy);background:var(--blue-bg)}
thead th.srt{color:var(--navy)}
.si{margin-left:3px;opacity:.5}
.srt .si{opacity:1}
tbody tr{border-bottom:1px solid var(--border);transition:background .1s}
tbody tr:last-child{border-bottom:none}
tbody tr:hover{background:#F8FAFC}
td{padding:9px 13px;vertical-align:middle}
.inst{font-weight:600;max-width:210px}
.gname{font-size:11px;color:var(--muted);margin-top:2px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;max-width:210px}

/* progress bar */
.pb-wrap{width:90px;background:var(--bg);border-radius:100px;height:6px;overflow:hidden}
.pb{height:100%;border-radius:100px;background:var(--blue);transition:width .8s cubic-bezier(.4,0,.2,1)}
.pb.red{background:var(--red)}
.pb.grn{background:var(--green)}
.pct-lbl{font-size:12px;font-weight:700;margin-top:3px}

/* quarterly icons */
.qi{display:flex;gap:3px;align-items:center}
.q{width:18px;height:18px;border-radius:50%;display:inline-flex;align-items:center;justify-content:center;font-size:10px;font-weight:800}
.q.ok{background:var(--green-bg);color:var(--green)}
.q.bad{background:var(--red-bg);color:var(--red)}
.q.fut{background:var(--bg);color:#CBD5E1}

/* badges */
.rbdg{display:inline-flex;align-items:center;padding:2px 8px;border-radius:100px;font-size:11px;font-weight:700;background:var(--blue-bg);color:var(--navy)}
.sbdg{display:inline-flex;align-items:center;padding:2px 8px;border-radius:100px;font-size:11px;font-weight:600}
.sbdg.cert{background:var(--green-bg);color:#065F46}
.sbdg.subm{background:var(--amber-bg);color:#92400E}
.sbdg.none{background:var(--bg);color:var(--muted)}
.sbdg.rs-submitted{background:#DBEAFE;color:#1E40AF;font-weight:700}
.sbdg.rs-pending{background:#FEF3C7;color:#92400E;font-weight:700}
.sbdg.rs-awaiting{background:#F3F4F6;color:#6B7280;font-weight:600}
.arf{display:inline-flex;align-items:center;justify-content:center;width:20px;height:20px;background:var(--red-bg);color:var(--red);border-radius:50%;font-size:11px}

/* ─── VIEW 2: PORTFOLIO ──────────────────────────────────── */
.rc-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:16px;margin-bottom:20px}
.rc{background:var(--card);border:2px solid var(--border);border-radius:var(--r);padding:20px;cursor:pointer;transition:all .2s;box-shadow:0 1px 4px rgba(0,0,0,.04)}
.rc:hover{border-color:var(--navy);transform:translateY(-2px);box-shadow:0 8px 24px rgba(27,58,143,.12)}
.rc.on{border-color:var(--navy);background:var(--blue-bg)}
.rc-lbl{font-size:12px;font-weight:800;color:var(--muted);text-transform:uppercase;letter-spacing:.5px;margin-bottom:8px}
.rc-budget{font-size:24px;font-weight:800;letter-spacing:-.5px;line-height:1;margin-bottom:3px}
.rc-sub{font-size:12px;color:var(--muted);margin-bottom:10px}
.rc-pct{font-size:14px;font-weight:700;color:var(--navy);margin-top:5px}
.rgl{background:var(--card);border-radius:var(--r);border:1px solid var(--border);overflow:hidden;box-shadow:0 1px 4px rgba(0,0,0,.04)}
.rgl-hdr{padding:11px 16px;background:var(--blue-bg);border-bottom:1px solid var(--blue-mid);font-size:13px;font-weight:700;color:var(--navy)}

/* ─── VIEW 3: MATRIX ─────────────────────────────────────── */
.mtx-wrap{background:var(--card);border-radius:var(--r);border:1px solid var(--border);overflow:auto;box-shadow:0 1px 4px rgba(0,0,0,.04);margin-bottom:24px}
.mtx th,.mtx td{padding:7px 10px;text-align:center;white-space:nowrap;border-bottom:1px solid var(--border)}
.mtx th:first-child,.mtx td:first-child{text-align:left;position:sticky;left:0;background:inherit;z-index:2;padding-left:14px;min-width:190px}
.mtx thead th{background:#F8FAFC;font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.5px;color:var(--muted);position:sticky;top:0;z-index:3}
.mtx thead th:first-child{z-index:4}
.mtx thead th.pd{background:#FFF5F5;color:var(--red)}
.mc{width:26px;height:26px;border-radius:6px;display:inline-flex;align-items:center;justify-content:center;font-size:13px;font-weight:700;margin:0 auto}
.mc.ok{background:var(--green-bg);color:var(--green)}
.mc.bad{background:var(--red-bg);color:var(--red)}
.mc.fut{background:transparent;color:#CBD5E1}
.mtx-foot td{font-weight:700;background:#F8FAFC;border-top:2px solid var(--border);font-size:12px;color:var(--muted)}
.sec-ttl{font-size:15px;font-weight:800;color:var(--text);margin-bottom:12px;letter-spacing:-.2px}

/* ─── VIEW 4: ALERTS ─────────────────────────────────────── */
.alrt-panel{background:#FEF2F2;border:1px solid #FECACA;border-radius:var(--r);margin-bottom:18px;overflow:hidden}
.alrt-hdr{display:flex;align-items:center;justify-content:space-between;padding:13px 18px;cursor:pointer;user-select:none}
.alrt-ttl{display:flex;align-items:center;gap:9px;font-size:14px;font-weight:700;color:#991B1B}
.alrt-n{background:var(--red);color:#fff;padding:1px 9px;border-radius:100px;font-size:12px;font-weight:800}
.alrt-tog{font-size:12px;color:#991B1B;font-weight:600}
.alrt-body{padding:0 18px 14px}
.alrt-item{display:flex;align-items:center;gap:10px;padding:9px 0;border-top:1px solid #FECACA;font-size:13px}
.alrt-inst{font-weight:700;min-width:180px}
.alrt-det{color:#7F1D1D;font-size:12px}
.sf-row{display:flex;gap:7px;margin-bottom:14px}
.sf-btn{background:var(--bg);border:1.5px solid var(--border);color:var(--muted);font-family:var(--font);font-size:12px;font-weight:600;padding:5px 13px;border-radius:8px;cursor:pointer;transition:all .15s}
.sf-btn.on{background:var(--navy);border-color:var(--navy);color:#fff}

/* ─── VIEW 5: CARDS ──────────────────────────────────────── */
.cards{display:grid;grid-template-columns:repeat(auto-fill,minmax(300px,1fr));gap:16px}
.gc{background:var(--card);border:1.5px solid var(--border);border-radius:var(--r);overflow:hidden;cursor:pointer;transition:all .2s;box-shadow:0 1px 4px rgba(0,0,0,.04)}
.gc:hover{box-shadow:0 8px 24px rgba(0,0,0,.1);transform:translateY(-2px)}
.gc.risk{border-color:#FECACA}
.gc.exp{border-color:var(--navy)}
.gc-hdr{padding:15px 17px 11px}
.gc-bdgs{display:flex;gap:5px;margin-bottom:7px;align-items:center;flex-wrap:wrap}
.gc-inst{font-size:15px;font-weight:800;line-height:1.3;margin-bottom:3px}
.gc-gname{font-size:11px;color:var(--muted);overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.gc-metrics{display:flex;gap:14px;padding:11px 17px;background:#F8FAFC;border-top:1px solid var(--border)}
.gc-m{display:flex;flex-direction:column;gap:1px}
.gc-ml{font-size:10px;text-transform:uppercase;letter-spacing:.5px;font-weight:700;color:var(--muted)}
.gc-mv{font-size:14px;font-weight:700}
.gc-qi{display:flex;align-items:center;gap:3px;padding:9px 17px;border-top:1px solid var(--border)}
.gc-ql{font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.5px;color:var(--muted);margin-right:4px}
.gc-det{overflow:hidden;max-height:0;transition:max-height .3s ease}
.gc-det.open{max-height:280px}
.gc-det-inner{padding:12px 17px;border-top:1px solid var(--border)}
.dr{display:flex;justify-content:space-between;padding:5px 0;border-bottom:1px solid var(--border);font-size:12px}
.dr:last-child{border-bottom:none}
.dl{color:var(--muted);font-weight:500}
.dv{font-weight:600;text-align:right;max-width:55%}

/* ─── MISC ───────────────────────────────────────────────── */
.empty{text-align:center;padding:48px 24px;color:var(--muted);font-size:14px}
.empty .ico{font-size:36px;margin-bottom:12px}
.num{font-variant-numeric:tabular-nums}
.inst-link{color:var(--navy);text-decoration:none;font-weight:600}
.inst-link:hover{text-decoration:underline;color:var(--navy-light)}

/* ─── VIEW 6: DEADLINE CALENDAR ─────────────────────────── */
.dl-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(680px,1fr));gap:20px}
.dl-card{background:var(--card);border-radius:var(--r);border:1px solid var(--border);overflow:hidden;box-shadow:0 1px 4px rgba(0,0,0,.05)}
.dl-card-top{height:4px}
.dl-hdr{display:flex;align-items:flex-start;justify-content:space-between;padding:16px 20px 12px;border-bottom:1px solid var(--border);flex-wrap:wrap;gap:10px}
.dl-round-name{font-size:18px;font-weight:800;letter-spacing:-.3px}
.dl-round-fy{font-size:12px;color:var(--muted);font-weight:500;margin-top:2px}
.dl-hdr-meta{display:flex;gap:18px;flex-wrap:wrap}
.dl-hdr-stat{display:flex;flex-direction:column;align-items:flex-end}
.dl-hdr-stat-val{font-size:15px;font-weight:700;color:var(--text)}
.dl-hdr-stat-lbl{font-size:10px;text-transform:uppercase;letter-spacing:.5px;font-weight:700;color:var(--muted)}
.dl-table{width:100%;border-collapse:collapse;font-size:13px}
.dl-table th{background:#F8FAFC;padding:9px 16px;text-align:left;font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.6px;color:var(--muted);border-bottom:1px solid var(--border)}
.dl-table td{padding:10px 16px;border-bottom:1px solid var(--border);vertical-align:middle}
.dl-table tr:last-child td{border-bottom:none}
.dl-table tr:hover td{background:#FAFBFC}
.dl-row-warn td{background:#FFFBF5}
.dl-row-warn:hover td{background:#FEF3C7}
.dl-row-final td{background:#F8FAFC;font-weight:600}
.dl-row-final-warn td{background:#FFF5F5}
.dl-period{font-weight:600;white-space:nowrap}
.dl-date{white-space:nowrap;font-variant-numeric:tabular-nums;color:var(--text)}
.dl-date-full{font-weight:600}
.dl-date-year{font-size:11px;color:var(--muted);display:block}
.dl-days{font-size:12px;font-weight:600;white-space:nowrap}
.dl-days.past{color:var(--muted)}
.dl-days.overdue{color:var(--red)}
.dl-days.soon{color:#1D4ED8}
.dl-days.future{color:var(--muted)}
.dl-bar-wrap{width:120px;background:var(--bg);border-radius:100px;height:5px;margin-top:4px}
.dl-bar{height:100%;border-radius:100px;transition:width .6s ease}
.dl-count{white-space:nowrap}
.dl-count-num{font-size:14px;font-weight:700}
.dl-count-den{font-size:12px;color:var(--muted)}
.dl-status{display:inline-flex;align-items:center;padding:3px 10px;border-radius:100px;font-size:12px;font-weight:700;white-space:nowrap}
.dl-status.done{background:var(--green-bg);color:var(--green)}
.dl-status.missing{background:var(--red-bg);color:var(--red)}
.dl-status.partial{background:var(--amber-bg);color:var(--amber)}
.dl-status.upcoming{background:var(--blue-bg);color:var(--navy)}
.dl-status.future{background:var(--bg);color:var(--muted)}
.dl-legend{display:flex;gap:16px;padding:10px 20px;border-top:1px solid var(--border);background:#F8FAFC;flex-wrap:wrap}
.dl-legend-item{display:flex;align-items:center;gap:5px;font-size:11px;font-weight:600;color:var(--muted)}
.dl-legend-dot{width:10px;height:10px;border-radius:50%}

/* ─── DEADLINE BAR ───────────────────────────────────────── */
.dbar{background:#0F172A;padding:7px 24px;display:flex;align-items:center;gap:16px;flex-wrap:wrap;border-bottom:2px solid #1E293B}
.dbar-sec{display:flex;align-items:center;gap:7px;flex-wrap:wrap}
.dbar-lbl{font-size:10px;font-weight:800;text-transform:uppercase;letter-spacing:.8px;white-space:nowrap}
.dbar-lbl.past-lbl{color:#F87171}
.dbar-lbl.up-lbl{color:#60A5FA}
.dbar-div{width:1px;height:20px;background:#334155;flex-shrink:0;margin:0 4px}
.dchip{display:inline-flex;align-items:center;gap:4px;padding:3px 9px;border-radius:100px;font-size:11px;font-weight:700;white-space:nowrap;line-height:1.3}
.dchip.overdue{background:#450A0A;color:#FCA5A5;border:1px solid #7F1D1B}
.dchip.recent{background:#422006;color:#FED7AA;border:1px solid #7C2D12}
.dchip.soon{background:#0C2A4A;color:#93C5FD;border:1px solid #1D4ED8}
.dchip.future{background:#1E293B;color:#94A3B8;border:1px solid #334155}
.dchip.done{background:#052E16;color:#6EE7B7;border:1px solid #065F46}

/* per-grant deadline chip in table */
.gdl{display:inline-flex;flex-direction:column;align-items:flex-start;gap:2px}
.gdl-chip{display:inline-flex;align-items:center;padding:2px 7px;border-radius:6px;font-size:11px;font-weight:700;white-space:nowrap}
.gdl-chip.overdue{background:var(--red-bg);color:var(--red)}
.gdl-chip.soon{background:#DBEAFE;color:#1D4ED8}
.gdl-chip.ok{background:var(--green-bg);color:var(--green)}
.gdl-chip.future{background:var(--bg);color:var(--muted)}
@media(max-width:900px){
  .kpi-row{grid-template-columns:repeat(2,1fr)}
  .rc-grid{grid-template-columns:repeat(2,1fr)}
}

/* ─── BURN RATE BADGES ───────────────────────────────────── */
.br-badge{display:inline-flex;align-items:center;padding:2px 8px;border-radius:100px;font-size:11px;font-weight:700;white-space:nowrap}
.br-ok{background:#D1FAE5;color:#065F46}
.br-warn{background:#FEF3C7;color:#92400E}
.br-risk{background:#FEE2E2;color:#991B1B}

/* ─── UNSPENT FUNDS AT RISK PANEL ───────────────────────── */
.unspent-panel{background:#FFF7ED;border:1px solid #FED7AA;border-radius:var(--r);margin-bottom:18px;overflow:hidden}
.unspent-hdr{display:flex;align-items:center;justify-content:space-between;padding:13px 18px;cursor:pointer;user-select:none}
.od-panel{border-radius:var(--r);margin-bottom:18px;overflow:hidden;border:1px solid #FECACA;background:#FEF2F2}
.od-hdr{display:flex;align-items:center;justify-content:space-between;padding:12px 18px;font-weight:700;font-size:13px;color:var(--red);cursor:pointer;user-select:none}
.od-hdr-ttl{display:flex;align-items:center;gap:8px}
.od-total{background:var(--red);color:#fff;padding:1px 9px;border-radius:100px;font-size:12px;font-weight:800}
.od-body{padding:0 18px 12px}
.od-group{margin-top:10px}
.od-group-hdr{display:flex;align-items:center;gap:8px;font-size:12px;font-weight:700;color:var(--red);margin-bottom:4px}
.od-group-hdr.amber{color:var(--amber)}
.od-cnt{background:var(--red-bg);color:var(--red);padding:0 7px;border-radius:100px;font-size:11px;font-weight:700}
.od-cnt.amber{background:var(--amber-bg);color:var(--amber)}
.od-row{display:flex;align-items:center;gap:8px;padding:5px 0;border-top:1px solid #FECACA;font-size:12px}
.od-row.amber{border-top-color:#FDE68A}
.od-gname{flex:1;font-weight:600;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;max-width:220px}
.od-detail{color:var(--red);font-weight:700;white-space:nowrap}
.od-detail.amber{color:var(--amber)}
.unspent-ttl{display:flex;align-items:center;gap:9px;font-size:14px;font-weight:700;color:#92400E}
.unspent-n{background:#F97316;color:#fff;padding:1px 9px;border-radius:100px;font-size:12px;font-weight:800}

/* ─── TOOL BUTTONS (export / print) ─────────────────────── */
.tool-btn{background:#fff;border:1.5px solid var(--border);color:var(--muted);font-family:var(--font);font-size:12px;font-weight:600;padding:5px 13px;border-radius:8px;cursor:pointer;transition:all .15s;white-space:nowrap}
.tool-btn:hover{border-color:var(--navy);color:var(--navy)}

/* ─── LOGO ───────────────────────────────────────────────── */
.hdr-logo{height:36px;width:auto;object-fit:contain;border-radius:4px}

/* ─── PRINT STYLES ───────────────────────────────────────── */
@media print {
  *{animation:none!important;transition:none!important;opacity:1!important}
  .hdr,.fbar,.dbar,#dbar,.tabs,.alrt-panel,.unspent-panel,.od-panel,.tool-btn,.pills,.srch,.rcount,.sf-row{display:none!important}
  .view{display:none!important}
  .view.on{display:block!important}
  body{font-size:12px;background:#fff}
  .main{padding:0;max-width:100%}
  .tbl-wrap{box-shadow:none;border:1px solid #ccc;overflow:visible!important}
  .kpi-row{grid-template-columns:repeat(4,1fr)}
  @page{margin:1.5cm}
}
</style>
</head>
<body>
<div id="pw-gate" style="display:flex;position:fixed;inset:0;z-index:9999;background:#1B3A8F;align-items:center;justify-content:center;font-family:'Plus Jakarta Sans',system-ui,sans-serif"><script>if(sessionStorage.getItem('swp_auth')==='1')document.currentScript.parentElement.style.display='none';</script>
  <div style="background:#0D1F5C;border-radius:14px;padding:40px 48px;width:100%;max-width:400px;box-shadow:0 20px 60px rgba(0,0,0,0.4);text-align:center">
    <div style="margin-bottom:20px"><img src="__LOGO_SRC__" alt="CVML Logo" style="width:220px;height:auto"></div>
    <div style="font-size:20px;font-weight:800;color:#fff;margin-bottom:4px">K12 SWP Grant Dashboard</div>
    <div style="font-size:13px;color:rgba(255,255,255,0.6);margin-bottom:28px">Central Mother Lode Region</div>
    <form id="pw-form">
      <input id="pw-input" type="password" placeholder="Enter password" autofocus
        style="width:100%;padding:12px 16px;border:2px solid rgba(255,255,255,0.2);border-radius:8px;font-size:15px;outline:none;margin-bottom:12px;font-family:inherit;background:rgba(255,255,255,0.1);color:#fff;">
      <div id="pw-err" style="display:none;color:#f87171;font-size:13px;margin-bottom:10px">Incorrect password. Please try again.</div>
      <button type="submit"
        style="width:100%;padding:12px;background:#00b3be;color:#fff;border:none;border-radius:8px;font-size:15px;font-weight:700;cursor:pointer;font-family:inherit">
        Sign In
      </button>
    </form>
  </div>
</div>
<div id="app">

<!-- HEADER -->
<header class="hdr">
  <div class="hdr-top">
    <div class="hdr-title">
      <img src="__LOGO_SRC__" alt="CVML Logo" class="hdr-logo" onerror="this.style.display='none'">
      K12 SWP Grant Dashboard
    </div>
    <div class="hdr-meta">Generated: <span id="gen-ts"></span></div>
  </div>
  <nav class="tabs">
    <button class="tab on" data-v="1">⚡ Command Center</button>
    <button class="tab" data-v="2">📊 Portfolio Overview</button>
    <button class="tab" data-v="3">🗓 Status Matrix</button>
    <button class="tab" data-v="4">⚠️ Alert Dashboard</button>
    <button class="tab" data-v="5">🃏 Grant Cards</button>
    <button class="tab" data-v="6">📋 Deadlines</button>
    <button class="tab" data-v="7">📈 Lifecycle</button>
  </nav>
  <div class="fbar">
    <div class="pills">
      <button class="pill on" data-r="all">All Rounds</button>
      <button class="pill" data-r="R6">R6</button>
      <button class="pill" data-r="R7">R7</button>
      <button class="pill" data-r="R8">R8</button>
    </div>
    <input type="search" class="srch" id="srch" placeholder="🔍  Search institution…">
    <button onclick="exportCSV()" class="tool-btn" title="Export current view to CSV">⬇ Export CSV</button>
    <button onclick="printDashboard()" class="tool-btn" title="Print or save as PDF">🖨 Print</button>
    <span class="rcount" id="rcount"></span>
  </div>
</header>

<!-- DEADLINE BAR -->
<div id="dbar" class="dbar"></div>

<!-- VIEWS -->
<main class="main">
  <div id="v1" class="view on"></div>
  <div id="v2" class="view"></div>
  <div id="v3" class="view"></div>
  <div id="v4" class="view"></div>
  <div id="v5" class="view"></div>
  <div id="v6" class="view"></div>
  <div id="v7" class="view"></div>
</main>

</div><!-- #app -->

<script>
// ─── DATA (injected by Python) ────────────────────────────────────────────
const DATA = __DATA_JSON__;
const TODAY = new Date('__TODAY__T00:00:00');

// ─── STATE ────────────────────────────────────────────────────────────────
const S = {
  view: 1,
  round: 'all',
  search: '',
  sortCol: 'institution',
  sortDir: 1,
  alertOpen: false,
  unspentOpen: false,
  odOpen: false,
  statusFilter: 'all',
  expandedCard: null,
  expandedRound: null,
};

// ─── HELPERS ──────────────────────────────────────────────────────────────
const $c = (v, dec=0) => v == null ? '—' :
  '$' + Math.round(v).toLocaleString('en-US');

const $p = n => Math.round(n * 100) + '%';
const $pf = n => (n * 100).toFixed(1) + '%';

function pastDue(ds) {
  return new Date(ds + 'T00:00:00') <= TODAY;
}

function pbHtml(pct, risk) {
  const w = Math.min(100, Math.round(pct * 100));
  const cls = risk ? 'red' : pct >= 0.75 ? 'grn' : '';
  return `<div class="pb-wrap"><div class="pb ${cls}" style="width:${w}%"></div></div>
          <div class="pct-lbl">${$p(pct)}</div>`;
}

function qiHtml(quarterly, dueDates) {
  return '<div class="qi">' + quarterly.map((v, i) => {
    const pd = pastDue(dueDates[i]);
    const [cls, sym] = v ? ['ok','✓'] : pd ? ['bad','✗'] : ['fut','·'];
    return `<span class="q ${cls}" title="${dueDates[i]}">${sym}</span>`;
  }).join('') + '</div>';
}

function finalIcon(done, dueDate, size='') {
  const st = size ? `style="width:${size};height:${size}"` : '';
  if (done) return `<span class="q ok" ${st}>✓</span>`;
  return pastDue(dueDate)
    ? `<span class="q bad" ${st}>✗</span>`
    : `<span class="q fut" ${st}>·</span>`;
}

function sbadge(status, certDate) {
  if (!status) return '';
  const s = status.toLowerCase();
  const cls = s.includes('certif') ? 'cert' : s.includes('submit') ? 'subm' : 'none';
  const dateTag = (s.includes('certif') && certDate)
    ? `<span style="font-size:10px;font-weight:400;opacity:0.85;margin-left:3px">${certDate}</span>`
    : '';
  return `<span class="sbdg ${cls}">${status}${dateTag}</span>`;
}

function rbadge(rk) { return `<span class="rbdg">${rk}</span>`; }
function arIcon() { return `<span class="arf" title="At risk — past-due report missing">⚠</span>`; }

function reportStatusBadge(g) {
  const s = (g.dashboardApproval || g.approvalStatus || '').toLowerCase();
  if (!s || s.includes('certif')) return '';
  const dateSuffix = g.submittedDate ? ` — ${g.submittedDate}` : '';
  if (s.includes('pending') || s === 'awaiting approval')
    return `<span class="sbdg rs-pending">⏳ Awaiting Approval${dateSuffix}</span>`;
  if (s.includes('submit'))
    return `<span class="sbdg rs-submitted">📋 Submitted${dateSuffix}</span>`;
  if (s.includes('awaiting'))
    return `<span class="sbdg rs-awaiting">⌛ Awaiting Submittal</span>`;
  return `<span class="sbdg none">${g.dashboardApproval || g.approvalStatus}</span>`;
}

function planIdTag(g) {
  return `<span style="font-size:11px;color:var(--muted);font-weight:400">Plan #${g.planId}</span>`;
}

function daysFrom(ds) {
  return Math.round((new Date(ds + 'T00:00:00') - TODAY) / 86400000);
}

function fmtDate(ds) {
  return new Date(ds + 'T00:00:00').toLocaleDateString('en-US', {month:'short', day:'numeric', year:'numeric'});
}

function novaUrl(g) {
  return `https://nova.cccco.edu/swpk/fiscal-reports/plans/${g.planId}?duration=${g.yearCode}`;
}

function instLink(g) {
  return `<a class="inst-link" href="${novaUrl(g)}" target="_blank" rel="noopener"
    title="View ${g.institution} grant on NOVA">${g.institution}</a>`;
}

// ─── BURN RATE HELPERS ────────────────────────────────────────────────────
// Compares actual spend % vs expected spend % based on lifecycle position
function burnRate(g, rd) {
  const start = new Date(rd.grantFYStart + 'T00:00:00');
  const end   = new Date(rd.finalReportDue + 'T00:00:00');
  const total = end - start;
  const elapsed = Math.max(0, Math.min(total, TODAY - start));
  const expectedPct = elapsed / total;
  const actualPct   = g.pctSpent;
  const delta       = actualPct - expectedPct;
  return { expectedPct, actualPct, delta };
}

function burnBadge(g, rd) {
  const { expectedPct, delta } = burnRate(g, rd);
  let cls, label, title;
  if (delta >= -0.05)       { cls = 'br-ok';   label = '▲ On Pace'; }
  else if (delta >= -0.15)  { cls = 'br-warn';  label = '▼ Slightly Behind'; }
  else                       { cls = 'br-risk';  label = '▼ Under-Spending'; }
  title = `Expected ${Math.round(expectedPct*100)}% spent by now`;
  return `<span class="br-badge ${cls}" title="${title}">${label}</span>`;
}

// ─── UNSPENT FUNDS AT RISK HELPER ─────────────────────────────────────────
function unspentRiskGrants() {
  // Grants in final year (≤365 days to final report) with >20% budget unspent and final report not done
  return allGrants().filter(g => {
    const daysLeft = daysFrom(g.finalReportDue);
    return daysLeft >= 0 && daysLeft <= 365 && !g.finalReport && g.pctSpent < 0.80;
  }).sort((a, b) => (a.budget - a.spent) < (b.budget - b.spent) ? 1 : -1);
}

function renderOverduePanel() {
  const grants = allGrants();

  // --- Group 1: Missed quarterly reports (one row per missed quarter per grant) ---
  const missedReports = [];
  for (const g of grants) {
    for (let i = 0; i < g.quarterlyDueDates.length; i++) {
      if (pastDue(g.quarterlyDueDates[i]) && !g.quarterly[i]) {
        missedReports.push({
          g,
          label: g.quarterlyLabels[i].split(' (')[0],
          date:  g.quarterlyDueDates[i],
          days:  Math.abs(daysFrom(g.quarterlyDueDates[i]))
        });
      }
    }
  }
  missedReports.sort((a, b) => daysFrom(a.date) - daysFrom(b.date));

  // --- Group 2: Missed final reports ---
  const missedFinal = grants
    .filter(g => pastDue(g.finalReportDue) && !g.finalReport)
    .map(g => ({ g, days: Math.abs(daysFrom(g.finalReportDue)) }))
    .sort((a, b) => daysFrom(a.g.finalReportDue) - daysFrom(b.g.finalReportDue));

  // --- Group 3: Behind on spending ---
  const behindSpend = [];
  for (const g of grants) {
    if (g.budget <= 0) continue;
    const rd = DATA.rounds[g.roundKey];
    if (!rd) continue;
    const { delta } = burnRate(g, rd);
    if (delta < -0.05) {
      behindSpend.push({ g, delta });
    }
  }
  behindSpend.sort((a, b) => a.delta - b.delta);

  const total = missedReports.length + missedFinal.length + behindSpend.length;
  if (total === 0) return '';  // Panel hidden when nothing is overdue

  function groupHtml(title, rowsHtml, isAmber) {
    if (!rowsHtml) return '';
    const cls = isAmber ? 'amber' : '';
    const cnt = isAmber ? behindSpend.length : (title.includes('Final') ? missedFinal.length : missedReports.length);
    return `<div class="od-group">
      <div class="od-group-hdr ${cls}">${title}<span class="od-cnt ${cls}">${cnt}</span></div>
      ${rowsHtml}
    </div>`;
  }

  const mrRows = missedReports.map(r =>
    `<div class="od-row">
      ${rbadge(r.g.roundKey)}
      <div class="od-gname" title="${r.g.grantName}">${r.g.grantName}</div>
      <div class="od-detail">${r.label} · ${r.days}d overdue</div>
    </div>`
  ).join('');

  const mfRows = missedFinal.map(r =>
    `<div class="od-row">
      ${rbadge(r.g.roundKey)}
      <div class="od-gname" title="${r.g.grantName}">${r.g.grantName}</div>
      <div class="od-detail">Final Report · ${r.days}d overdue</div>
    </div>`
  ).join('');

  const bsRows = behindSpend.map(r =>
    `<div class="od-row amber">
      ${rbadge(r.g.roundKey)}
      <div class="od-gname" title="${r.g.grantName}">${r.g.grantName}</div>
      <div class="od-detail amber">${Math.round(Math.abs(r.delta) * 100)}% behind pace</div>
    </div>`
  ).join('');

  const body = S.odOpen ? `<div class="od-body">
      ${groupHtml('Missed Reports (Q2/Q4)', mrRows, false)}
      ${groupHtml('Missed Final Reports', mfRows, false)}
      ${groupHtml('Behind on Spending', bsRows, true)}
    </div>` : '';

  return `<div class="od-panel">
    <div class="od-hdr" onclick="togOD()">
      <div class="od-hdr-ttl">⚠ Overdue &amp; At-Risk<span class="od-total">${total}</span></div>
      <span class="alrt-tog">${S.odOpen ? 'Collapse ▲' : 'Expand ▼'}</span>
    </div>
    ${body}
  </div>`;
}

// ─── EXPORT / PRINT HELPERS ───────────────────────────────────────────────
function exportCSV() {
  const gs = filterG(allGrants());
  const headers = ['Round','Institution','Plan ID','Grant Name','Budget','Spent','% Spent','At Risk','Plan Status'];
  const rows = gs.map(g => [
    g.roundKey, `"${g.institution}"`, g.planId, `"${g.grantName}"`,
    g.budget, g.spent, Math.round(g.pctSpent*100)+'%',
    g.atRisk ? 'Yes' : 'No', `"${g.planStatus||g.approvalStatus||''}"`
  ]);
  const csv = [headers, ...rows].map(r => r.join(',')).join('\n');
  const blob = new Blob([csv], {type:'text/csv'});
  const url  = URL.createObjectURL(blob);
  const a    = document.createElement('a');
  a.href = url; a.download = 'k12-swp-grants.csv'; a.click();
  URL.revokeObjectURL(url);
}

function printDashboard() {
  render();
  requestAnimationFrame(() => requestAnimationFrame(() => window.print()));
}

// ─── DEADLINE HELPERS ─────────────────────────────────────────────────────
// Returns a small chip showing the grant's most urgent deadline situation.
function grantDeadlineChip(g) {
  // 1. First overdue incomplete quarter
  for (let i = 0; i < g.quarterly.length; i++) {
    if (!g.quarterly[i] && pastDue(g.quarterlyDueDates[i])) {
      const d = Math.abs(daysFrom(g.quarterlyDueDates[i]));
      const lbl = g.quarterlyLabels[i].split(' (')[0];
      return `<span class="gdl-chip overdue">⚠ ${lbl}<br>${d}d overdue</span>`;
    }
  }
  // 2. Final report overdue
  if (!g.finalReport && pastDue(g.finalReportDue)) {
    const d = Math.abs(daysFrom(g.finalReportDue));
    return `<span class="gdl-chip overdue">⚠ Final Report<br>${d}d overdue</span>`;
  }
  // 3. Next upcoming quarter not yet submitted
  for (let i = 0; i < g.quarterly.length; i++) {
    if (!g.quarterly[i]) {
      const d = daysFrom(g.quarterlyDueDates[i]);
      const lbl = g.quarterlyLabels[i].split(' (')[0];
      const cls = d <= 45 ? 'soon' : 'future';
      return `<span class="gdl-chip ${cls}">${lbl}<br>due in ${d}d</span>`;
    }
  }
  // 4. Final report upcoming
  if (!g.finalReport) {
    const d = daysFrom(g.finalReportDue);
    return `<span class="gdl-chip ${d <= 60 ? 'soon' : 'future'}">Final Report<br>due in ${d}d</span>`;
  }
  return `<span class="gdl-chip ok">✓ All done</span>`;
}

// Collects all round-level deadlines, deduped by date+label, sorted by date.
function allDeadlines() {
  const map = {};  // key = dateStr → [{rk, label}]
  Object.entries(DATA.rounds).forEach(([rk, rd]) => {
    rd.quarterlyLabels.forEach((lbl, i) => {
      const ds = rd.quarterlyDueDates[i];
      const short = lbl.split(' (')[0];
      const key = ds + '|' + short;
      if (!map[key]) map[key] = {ds, label: short, rounds: [], isFinal: false};
      map[key].rounds.push(rk);
    });
    const key = rd.finalReportDue + '|Final Report';
    if (!map[key]) map[key] = {ds: rd.finalReportDue, label: 'Final Report', rounds: [], isFinal: true};
    map[key].rounds.push(rk);
  });
  return Object.values(map).sort((a, b) => a.ds.localeCompare(b.ds));
}

function renderDeadlineBar() {
  const dl = allDeadlines();
  const past   = dl.filter(d => daysFrom(d.ds) < 0 && daysFrom(d.ds) >= -75).reverse(); // last 75 days
  const coming = dl.filter(d => daysFrom(d.ds) >= 0).slice(0, 5);  // next 5

  function chip(d) {
    const days   = daysFrom(d.ds);
    const rks    = [...new Set(d.rounds)].join('/');
    const name   = `${rks} ${d.label}`;
    const dateShort = new Date(d.ds + 'T00:00:00')
      .toLocaleDateString('en-US', {month:'short', day:'numeric', year:'numeric'});
    if (days < 0) {
      const cls = days >= -14 ? 'recent' : 'overdue';
      return `<span class="dchip ${cls}">${name} · ${dateShort} · ${Math.abs(days)}d ago</span>`;
    }
    const cls = days === 0 ? 'recent' : days <= 30 ? 'soon' : 'future';
    return `<span class="dchip ${cls}">${name} · ${dateShort} · ${days === 0 ? 'TODAY' : 'in ' + days + 'd'}</span>`;
  }

  let html = '';
  if (past.length) {
    html += `<div class="dbar-sec"><span class="dbar-lbl past-lbl">⚠ Recently passed</span>
      ${past.map(chip).join('')}</div>`;
  }
  if (past.length && coming.length) {
    html += '<div class="dbar-div"></div>';
  }
  if (coming.length) {
    html += `<div class="dbar-sec"><span class="dbar-lbl up-lbl">📅 Upcoming</span>
      ${coming.map(chip).join('')}</div>`;
  }
  if (!past.length && !coming.length) {
    html = '<span style="font-size:12px;color:#475569">No deadlines in the next 6 months.</span>';
  }
  document.getElementById('dbar').innerHTML = html;
}

// ─── DATA ACCESS ──────────────────────────────────────────────────────────
function rounds() {
  const rk = S.round;
  const entries = Object.entries(DATA.rounds);
  return rk === 'all' ? entries : entries.filter(([k]) => k === rk);
}

function allGrants() {
  return rounds().flatMap(([rk, rd]) =>
    rd.grants.map(g => ({
      ...g, roundKey: rk, roundLabel: rd.label,
      yearCode: rd.yearCode,
      quarterlyDueDates: rd.quarterlyDueDates,
      quarterlyLabels: rd.quarterlyLabels,
      finalReportDue: rd.finalReportDue,
    }))
  );
}

function filterG(gs) {
  const q = S.search.toLowerCase().trim();
  if (!q) return gs;
  return gs.filter(g =>
    g.institution.toLowerCase().includes(q) ||
    g.grantName.toLowerCase().includes(q) ||
    g.planId.includes(q)
  );
}

function sortG(gs) {
  const { sortCol: col, sortDir: dir } = S;
  return [...gs].sort((a, b) => {
    let av = a[col], bv = b[col];
    if (typeof av === 'string') av = av.toLowerCase();
    if (typeof bv === 'string') bv = bv.toLowerCase();
    if (av < bv) return -dir;
    if (av > bv) return dir;
    return 0;
  });
}

function kpiStats(gs) {
  return {
    count: gs.length,
    budget: gs.reduce((s, g) => s + g.budget, 0),
    spent:  gs.reduce((s, g) => s + g.spent, 0),
    atRisk: gs.filter(g => g.atRisk).length,
  };
}

function setCount(n) {
  const el = document.getElementById('rcount');
  if (el) el.textContent = n === 1 ? '1 grant' : `${n} grants`;
}

function sa(col) {
  if (S.sortCol !== col) return `<span class="si">⇅</span>`;
  return `<span class="si">${S.sortDir === 1 ? '↑' : '↓'}</span>`;
}

// ─── VIEW 1: COMMAND CENTER ───────────────────────────────────────────────
function togUnspent() { S.unspentOpen = !S.unspentOpen; rv1(); }
function togOD()      { S.odOpen = !S.odOpen; rv1(); rv2(); }

function rv1() {
  const gs = filterG(allGrants());
  const sorted = sortG(gs);
  const st = kpiStats(gs);
  setCount(sorted.length);

  // Unspent Funds at Risk panel
  const riskGrants = unspentRiskGrants();
  let unspentPanel = '';
  if (riskGrants.length > 0) {
    let unspentBody = '';
    if (S.unspentOpen) {
      const unspentRows = riskGrants.map(g => {
        const unspent = g.budget - g.spent;
        const pctRemaining = g.budget > 0 ? Math.round((1 - g.pctSpent) * 100) : 0;
        const dl = daysFrom(g.finalReportDue);
        return `<div class="alrt-item">
          ${rbadge(g.roundKey)}
          <div class="alrt-inst">${instLink(g)} ${planIdTag(g)}</div>
          <div class="alrt-det">${$c(unspent)} unspent (${pctRemaining}% remaining) · ${dl}d until final report</div>
        </div>`;
      }).join('');
      unspentBody = `<div class="alrt-body">${unspentRows}</div>`;
    }
    unspentPanel = `<div class="unspent-panel">
      <div class="unspent-hdr" onclick="togUnspent()">
        <div class="unspent-ttl">
          <span>💰</span><span>Unspent Funds at Risk of Lapsing</span>
          <span class="unspent-n">${riskGrants.length}</span>
        </div>
        <span class="alrt-tog">${S.unspentOpen ? 'Collapse ▲' : 'Expand ▼'}</span>
      </div>
      ${unspentBody}
    </div>`;
  }

  const kpi = `<div class="kpi-row">
    <div class="kpi nvy">
      <div class="kpi-lbl">Total Budget</div>
      <div class="kpi-val num">${$c(st.budget)}</div>
      <div class="kpi-sub">${st.count} grants · Rounds 6 – 8</div>
    </div>
    <div class="kpi">
      <div class="kpi-lbl">Total Spent</div>
      <div class="kpi-val num">${$c(st.spent)}</div>
      <div class="kpi-sub">${st.budget > 0 ? $pf(st.spent/st.budget) : '0%'} of budget used</div>
    </div>
    <div class="kpi">
      <div class="kpi-lbl">Active Grants</div>
      <div class="kpi-val">${st.count}</div>
      <div class="kpi-sub">Rounds 5 – 8</div>
    </div>
    <div class="kpi ${st.atRisk > 0 ? 'red' : 'grn'}">
      <div class="kpi-lbl">At-Risk Grants</div>
      <div class="kpi-val">${st.atRisk}</div>
      <div class="kpi-sub">${st.atRisk > 0 ? 'Missing past-due reports' : 'All reports on track'}</div>
    </div>
  </div>`;

  let rows = sorted.length ? sorted.map(g =>
    `<tr>
      <td><div class="inst">${instLink(g)} ${planIdTag(g)}</div><div class="gname" title="${g.grantName}">${g.grantName}</div></td>
      <td>${rbadge(g.roundKey)}</td>
      <td class="num">${$c(g.budget)}</td>
      <td class="num">${$c(g.spent)}</td>
      <td>${pbHtml(g.pctSpent, g.atRisk)}</td>
      <td>${burnBadge(g, DATA.rounds[g.roundKey])}</td>
      <td>${qiHtml(g.quarterly, g.quarterlyDueDates)}</td>
      <td style="text-align:center">${finalIcon(g.finalReport, g.finalReportDue)}</td>
      <td>${grantDeadlineChip(g)}</td>
      <td>${sbadge(g.planStatus || g.approvalStatus, g.certifiedDate)}</td>
      <td>${g.atRisk ? arIcon() : ''}</td>
    </tr>`
  ).join('') : `<tr><td colspan="11"><div class="empty"><div class="ico">🔍</div>No grants match your search.</div></td></tr>`;

  const tbl = `<div class="tbl-wrap"><table>
    <thead><tr>
      <th onclick="ds('institution')" class="${S.sortCol==='institution'?'srt':''}">Institution ${sa('institution')}</th>
      <th>Round</th>
      <th onclick="ds('budget')" class="${S.sortCol==='budget'?'srt':''}">Budget ${sa('budget')}</th>
      <th onclick="ds('spent')" class="${S.sortCol==='spent'?'srt':''}">Spent ${sa('spent')}</th>
      <th onclick="ds('pctSpent')" class="${S.sortCol==='pctSpent'?'srt':''}">% Spent ${sa('pctSpent')}</th>
      <th>Burn Rate</th>
      <th>Quarterly Status</th>
      <th>Final</th>
      <th>Next Deadline</th>
      <th>Plan Status</th>
      <th></th>
    </tr></thead>
    <tbody>${rows}</tbody>
  </table></div>`;

  document.getElementById('v1').innerHTML = renderOverduePanel() + unspentPanel + kpi + tbl;
  animateBars('#v1');
}

// ─── VIEW 2: PORTFOLIO OVERVIEW ───────────────────────────────────────────
function rv2() {
  const q = S.search.toLowerCase().trim();
  const rds = rounds();

  const cards = rds.map(([rk, rd]) => {
    const t = rd.totals;
    const qSummary = t.quarterlyCounts.map((c, i) => {
      const lbl = rd.quarterlyLabels[i] ? rd.quarterlyLabels[i].split(' (')[0] : `Q${i+1}`;
      return `${c}/${t.grantCount}`;
    }).join(' · ');
    return `<div class="rc ${S.expandedRound===rk?'on':''}" onclick="togRound('${rk}')">
      <div class="rc-lbl">${rd.label} · FY ${rd.grantFY}</div>
      <div class="rc-budget num">${$c(t.budget)}</div>
      <div class="rc-sub">${t.grantCount} grants</div>
      <div class="pb-wrap" style="width:100%;height:8px;margin-bottom:4px">
        <div class="pb ${t.pctSpent>=.75?'grn':''}" style="width:${Math.min(100,Math.round(t.pctSpent*100))}%"></div>
      </div>
      <div class="rc-pct">${$pf(t.pctSpent)} spent · ${$c(t.spent)}</div>
      <div style="margin-top:7px;font-size:11px;color:var(--muted)">Quarterly: ${qSummary}</div>
      ${t.atRiskCount > 0 ? `<div style="margin-top:6px;font-size:11px;color:var(--red);font-weight:700">⚠ ${t.atRiskCount} at risk</div>` : `<div style="margin-top:6px;font-size:11px;color:var(--green);font-weight:600">✓ All on track</div>`}
    </div>`;
  }).join('');

  let listHtml = '';
  if (S.expandedRound) {
    const found = rds.find(([k]) => k === S.expandedRound);
    if (found) {
      const [rk, rd] = found;
      let gs = rd.grants.map(g => ({...g, yearCode: rd.yearCode}));
      if (q) gs = gs.filter(g =>
        g.institution.toLowerCase().includes(q) || g.grantName.toLowerCase().includes(q)
      );
      setCount(gs.length);
      const rows = gs.length ? gs.map(g =>
        `<tr>
          <td><div class="inst">${instLink(g)} ${planIdTag(g)}</div><div class="gname" title="${g.grantName}">${g.grantName}</div></td>
          <td class="num">${$c(g.budget)}</td>
          <td class="num">${$c(g.spent)}</td>
          <td>${pbHtml(g.pctSpent, g.atRisk)}</td>
          <td>${qiHtml(g.quarterly, rd.quarterlyDueDates)}</td>
          <td style="text-align:center">${finalIcon(g.finalReport, rd.finalReportDue)}</td>
          <td>${sbadge(g.planStatus || g.approvalStatus, g.certifiedDate)}</td>
          <td>${g.atRisk ? arIcon() : ''}</td>
        </tr>`
      ).join('') : `<tr><td colspan="8"><div class="empty"><div class="ico">🔍</div>No grants match.</div></td></tr>`;

      listHtml = `<div class="rgl">
        <div class="rgl-hdr">${rd.label} — ${gs.length} grant${gs.length!==1?'s':''} shown</div>
        <div class="tbl-wrap" style="border:none;border-radius:0;box-shadow:none"><table>
          <thead><tr>
            <th>Institution</th><th>Budget</th><th>Spent</th>
            <th>% Spent</th><th>Quarterly</th><th>Final</th><th>Status</th><th></th>
          </tr></thead>
          <tbody>${rows}</tbody>
        </table></div>
      </div>`;
    }
  } else {
    setCount(rds.reduce((s,[,rd])=>s+rd.totals.grantCount,0));
  }

  document.getElementById('v2').innerHTML =
    renderOverduePanel() + `<div class="rc-grid">${cards}</div>${listHtml}`;
  animateBars('#v2');
}

// ─── VIEW 3: STATUS MATRIX ────────────────────────────────────────────────
function rv3() {
  const q = S.search.toLowerCase().trim();
  const rds = rounds();
  let html = '';
  let total = 0;

  rds.forEach(([rk, rd]) => {
    let gs = rd.grants.map(g => ({...g, yearCode: rd.yearCode}));
    if (q) gs = gs.filter(g => g.institution.toLowerCase().includes(q) || g.grantName.toLowerCase().includes(q));
    total += gs.length;

    const allLabels = [...rd.quarterlyLabels, 'Final Report'];
    const allDates  = [...rd.quarterlyDueDates, rd.finalReportDue];
    const counts    = new Array(allLabels.length).fill(0);

    const hdrCols = allLabels.map((lbl, i) => {
      const pd = pastDue(allDates[i]);
      const short = lbl.split(' (')[0];
      const dateStr = allDates[i].slice(5).replace('-','/');
      return `<th class="${pd?'pd':''}">${short}<br><small style="font-weight:400;opacity:.7">${dateStr}</small></th>`;
    }).join('');

    const bodyRows = gs.map(g => {
      const qCols = g.quarterly.map((v, i) => {
        const pd = pastDue(allDates[i]);
        const [cls, sym] = v ? ['ok','✓'] : pd ? ['bad','✗'] : ['fut','—'];
        if (v) counts[i]++;
        return `<td><div class="mc ${cls}">${sym}</div></td>`;
      });
      // Final report
      const fi = allLabels.length - 1;
      const pd = pastDue(rd.finalReportDue);
      const [fcls, fsym] = g.finalReport ? ['ok','✓'] : pd ? ['bad','✗'] : ['fut','—'];
      if (g.finalReport) counts[fi]++;
      return `<tr><td><strong style="font-size:13px">${instLink(g)}</strong> ${planIdTag(g)}${g.atRisk?` <span class="arf" style="width:16px;height:16px;font-size:9px">⚠</span>`:''}</td>${qCols.join('')}<td><div class="mc ${fcls}">${fsym}</div></td></tr>`;
    }).join('');

    const footCols = counts.map((c, i) =>
      `<td><strong>${c}/${gs.length}</strong></td>`
    ).join('');

    html += `<div class="sec-ttl">${rd.label} — Submission Matrix</div>
    <div class="mtx-wrap">
      <table class="mtx">
        <thead><tr><th>Institution</th>${hdrCols}</tr></thead>
        <tbody>${bodyRows}</tbody>
        <tfoot><tr class="mtx-foot"><td><strong>Submitted</strong></td>${footCols}</tr></tfoot>
      </table>
    </div>`;
  });

  setCount(total);
  document.getElementById('v3').innerHTML = html ||
    '<div class="empty"><div class="ico">📋</div>No data for selected round.</div>';
}

// ─── VIEW 4: ALERT DASHBOARD ──────────────────────────────────────────────
function rv4() {
  const gs = filterG(allGrants());
  const atRisk = gs.filter(g => g.atRisk);
  const sorted = sortG(S.statusFilter === 'risk' ? atRisk :
    S.statusFilter === 'ok' ? gs.filter(g => !g.atRisk) : gs);
  setCount(sorted.length);

  // Alert panel
  let alertItems = '';
  if (S.alertOpen) {
    if (atRisk.length) {
      alertItems = '<div class="alrt-body">' + atRisk.map(g => {
        const missing = [
          ...g.quarterly.map((v, i) =>
            !v && pastDue(g.quarterlyDueDates[i])
              ? `${g.quarterlyLabels[i].split(' (')[0]} (${Math.abs(daysFrom(g.quarterlyDueDates[i]))}d overdue)`
              : null
          ).filter(Boolean),
          !g.finalReport && pastDue(g.finalReportDue)
            ? `Final Report (${Math.abs(daysFrom(g.finalReportDue))}d overdue)`
            : null,
        ].filter(Boolean);
        return `<div class="alrt-item">
          ${rbadge(g.roundKey)}
          <div class="alrt-inst">${instLink(g)} ${planIdTag(g)}</div>
          <div class="alrt-det">Missing: ${missing.join(' · ')}</div>
        </div>`;
      }).join('') + '</div>';
    } else {
      alertItems = '<div style="padding:14px 18px;color:#065F46;font-weight:600;font-size:13px">✅ All grants are on track — no missing past-due reports.</div>';
    }
  }

  const alertPanel = `<div class="alrt-panel">
    <div class="alrt-hdr" onclick="togAlert()">
      <div class="alrt-ttl">
        <span>⚠</span><span>Grants Needing Attention</span>
        <span class="alrt-n">${atRisk.length}</span>
      </div>
      <span class="alrt-tog">${S.alertOpen ? 'Collapse ▲' : 'Expand ▼'}</span>
    </div>
    ${alertItems}
  </div>`;

  const sfRow = `<div class="sf-row">
    <button class="sf-btn ${S.statusFilter==='all'?'on':''}" onclick="setSF('all')">All (${gs.length})</button>
    <button class="sf-btn ${S.statusFilter==='risk'?'on':''}" onclick="setSF('risk')">At-Risk (${atRisk.length})</button>
    <button class="sf-btn ${S.statusFilter==='ok'?'on':''}" onclick="setSF('ok')">On Track (${gs.length-atRisk.length})</button>
  </div>`;

  const rows = sorted.length ? sorted.map(g =>
    `<tr style="${g.atRisk?'background:#FFFAFA':''}">
      <td>
        <div class="inst">${instLink(g)}</div>
        <div style="font-size:11px;color:var(--muted);margin-top:1px">${planIdTag(g)}</div>
        ${g.atRisk ? `<div style="font-size:11px;color:var(--red);font-weight:600;margin-top:2px">⚠ Past-due report missing</div>` : ''}
      </td>
      <td>${rbadge(g.roundKey)}</td>
      <td class="num">${$c(g.budget)}</td>
      <td>${pbHtml(g.pctSpent, g.atRisk)}</td>
      <td>${qiHtml(g.quarterly, g.quarterlyDueDates)}</td>
      <td>${reportStatusBadge(g)}</td>
      <td style="text-align:center">${finalIcon(g.finalReport, g.finalReportDue)}</td>
      <td>${grantDeadlineChip(g)}</td>
      <td>${sbadge(g.planStatus || g.approvalStatus, g.certifiedDate)}</td>
    </tr>`
  ).join('') : `<tr><td colspan="9"><div class="empty"><div class="ico">✅</div>No grants in this category.</div></td></tr>`;

  const tbl = `<div class="tbl-wrap"><table>
    <thead><tr>
      <th onclick="ds('institution')" class="${S.sortCol==='institution'?'srt':''}">Institution ${sa('institution')}</th>
      <th>Round</th>
      <th onclick="ds('budget')" class="${S.sortCol==='budget'?'srt':''}">Budget ${sa('budget')}</th>
      <th onclick="ds('pctSpent')" class="${S.sortCol==='pctSpent'?'srt':''}">% Spent ${sa('pctSpent')}</th>
      <th>Quarterly Status</th>
      <th>Report Status</th>
      <th>Final Report</th>
      <th>Next Deadline</th>
      <th>Plan Status</th>
    </tr></thead>
    <tbody>${rows}</tbody>
  </table></div>`;

  document.getElementById('v4').innerHTML = alertPanel + sfRow + tbl;
  animateBars('#v4');
}

// ─── VIEW 5: GRANT CARDS ──────────────────────────────────────────────────
function rv5() {
  const gs = [...filterG(allGrants())].sort((a,b) =>
    a.institution.localeCompare(b.institution)
  );
  setCount(gs.length);

  if (!gs.length) {
    document.getElementById('v5').innerHTML =
      '<div class="empty"><div class="ico">🔍</div>No grants match your search.</div>';
    return;
  }

  const html = '<div class="cards">' + gs.map(g => {
    const exp = S.expandedCard === g.planId;
    const pctColor = g.atRisk ? 'var(--red)' : g.pctSpent >= 0.75 ? 'var(--green)' : 'var(--navy)';
    return `<div class="gc ${g.atRisk?'risk':''} ${exp?'exp':''}" onclick="togCard('${g.planId}')">
      <div class="gc-hdr">
        <div class="gc-bdgs">
          ${rbadge(g.roundKey)}
          ${sbadge(g.planStatus || g.approvalStatus, g.certifiedDate)}
          ${reportStatusBadge(g)}
          ${g.atRisk ? arIcon() : ''}
        </div>
        <div class="gc-inst">${instLink(g)} ${planIdTag(g)}</div>
        <div class="gc-gname" title="${g.grantName}">${g.grantName || '—'}</div>
      </div>
      <div class="gc-metrics">
        <div class="gc-m">
          <div class="gc-ml">Budget</div>
          <div class="gc-mv num">${$c(g.budget)}</div>
        </div>
        <div class="gc-m">
          <div class="gc-ml">Spent</div>
          <div class="gc-mv num">${$c(g.spent)}</div>
        </div>
        <div class="gc-m">
          <div class="gc-ml">% Spent</div>
          <div class="gc-mv" style="color:${pctColor}">${$p(g.pctSpent)}</div>
        </div>
      </div>
      <div style="padding:6px 17px 2px">
        <div class="pb-wrap" style="width:100%;height:5px">
          <div class="pb ${g.atRisk?'red':g.pctSpent>=.75?'grn':''}" style="width:${Math.min(100,Math.round(g.pctSpent*100))}%"></div>
        </div>
      </div>
      <div class="gc-qi">
        <span class="gc-ql">Quarterly</span>
        ${qiHtml(g.quarterly, g.quarterlyDueDates)}
        <span class="gc-ql" style="margin-left:6px">Final</span>
        ${finalIcon(g.finalReport, g.finalReportDue, '18px')}
      </div>
      <div style="padding:4px 17px 10px;display:flex;align-items:center;gap:6px">
        <span style="font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.5px;color:var(--muted)">Pace</span>
        ${burnBadge(g, DATA.rounds[g.roundKey])}
      </div>
      <div class="gc-det ${exp?'open':''}">
        <div class="gc-det-inner">
          <div class="dr"><span class="dl">Proposal ID</span><span class="dv">${g.planId}</span></div>
          <div class="dr"><span class="dl">Grant Name</span><span class="dv" title="${g.grantName}" style="overflow:hidden;text-overflow:ellipsis;white-space:nowrap">${g.grantName||'—'}</span></div>
          <div class="dr"><span class="dl">Lead LEA</span><span class="dv">${g.leadLEA||'—'}</span></div>
          <div class="dr"><span class="dl">Reporting Institutions</span><span class="dv">${g.instCount}</span></div>
          <div class="dr"><span class="dl">Approval Status</span><span class="dv">${g.dashboardApproval||g.approvalStatus||'—'}</span></div>
          <div class="dr"><span class="dl">Budget Remaining</span><span class="dv num">${$c(g.budget - g.spent)}</span></div>
          <div class="dr"><span class="dl">Final Report Due</span><span class="dv">${g.finalReportDue}</span></div>
        </div>
      </div>
    </div>`;
  }).join('') + '</div>';

  document.getElementById('v5').innerHTML = html;
  animateBars('#v5');
}

// ─── VIEW 6: DEADLINES ────────────────────────────────────────────────────
const ROUND_ACCENT = {R5:'#6D28D9', R6:'#0369A1', R7:'#047857', R8:'#B45309'};

function rv6() {
  const rds = rounds();
  if (!rds.length) {
    document.getElementById('v6').innerHTML =
      '<div class="empty"><div class="ico">📋</div>No rounds selected.</div>';
    return;
  }

  function periodRow(label, ds, submitted, total, isFinal) {
    const days    = daysFrom(ds);
    const pd      = pastDue(ds);
    const pct     = total > 0 ? submitted / total : 0;
    const allDone = submitted === total;
    const noneDone= submitted === 0;

    // days label
    let daysHtml, daysCls;
    if (!pd) {
      daysCls = days <= 45 ? 'soon' : 'future';
      daysHtml = `in ${days} day${days!==1?'s':''}`;
    } else {
      daysCls = allDone ? 'past' : 'overdue';
      daysHtml = `${Math.abs(days)} day${Math.abs(days)!==1?'s':''} ago`;
    }

    // status chip
    let statusHtml;
    if (!pd) {
      statusHtml = days <= 45
        ? `<span class="dl-status upcoming">🔔 Due in ${days}d</span>`
        : `<span class="dl-status future">Upcoming</span>`;
    } else if (allDone) {
      statusHtml = `<span class="dl-status done">✓ All ${total} submitted</span>`;
    } else if (noneDone) {
      statusHtml = `<span class="dl-status missing">✗ None submitted</span>`;
    } else {
      statusHtml = `<span class="dl-status ${submitted/total>=.75?'partial':'missing'}">⚠ ${submitted}/${total} submitted</span>`;
    }

    // progress bar colour
    const barColor = !pd ? '#3B82F6' : allDone ? '#059669' : submitted/total >= .5 ? '#D97706' : '#DC2626';

    const rowCls = isFinal
      ? (pd && !allDone ? 'dl-row-final-warn' : 'dl-row-final')
      : (pd && !allDone ? 'dl-row-warn' : '');

    return `<tr class="${rowCls}">
      <td class="dl-period">${isFinal ? '🏁 ' : ''}${label}</td>
      <td class="dl-date">
        <span class="dl-date-full">${fmtDate(ds)}</span>
      </td>
      <td class="dl-days ${daysCls}">${daysHtml}</td>
      <td class="dl-count">
        <span class="dl-count-num">${submitted}</span><span class="dl-count-den"> / ${total}</span>
        <div class="dl-bar-wrap">
          <div class="dl-bar" style="width:${Math.round(pct*100)}%;background:${barColor}"></div>
        </div>
      </td>
      <td>${statusHtml}</td>
    </tr>`;
  }

  const cards = rds.map(([rk, rd]) => {
    const t      = rd.totals;
    const accent = ROUND_ACCENT[rk] || 'var(--navy)';
    const rows   = rd.quarterlyLabels.map((lbl, i) =>
      periodRow(lbl.split(' (')[0], rd.quarterlyDueDates[i], t.quarterlyCounts[i], t.grantCount, false)
    );
    rows.push(periodRow('Final Report', rd.finalReportDue, t.finalCount, t.grantCount, true));

    return `<div class="dl-card">
      <div class="dl-card-top" style="background:${accent}"></div>
      <div class="dl-hdr">
        <div>
          <div class="dl-round-name" style="color:${accent}">${rd.label}</div>
          <div class="dl-round-fy">Grant Award FY ${rd.grantFY} &nbsp;·&nbsp; Final Report Due ${fmtDate(rd.finalReportDue)}</div>
        </div>
        <div class="dl-hdr-meta">
          <div class="dl-hdr-stat">
            <span class="dl-hdr-stat-val">${t.grantCount}</span>
            <span class="dl-hdr-stat-lbl">Grants</span>
          </div>
          <div class="dl-hdr-stat">
            <span class="dl-hdr-stat-val num">${$c(t.budget)}</span>
            <span class="dl-hdr-stat-lbl">Total Budget</span>
          </div>
          <div class="dl-hdr-stat">
            <span class="dl-hdr-stat-val num">${$p(t.pctSpent)}</span>
            <span class="dl-hdr-stat-lbl">% Spent</span>
          </div>
          <div class="dl-hdr-stat">
            <span class="dl-hdr-stat-val ${t.atRiskCount > 0 ? 'red' : ''}" style="${t.atRiskCount > 0 ? 'color:var(--red)' : 'color:var(--green)'}">${t.atRiskCount > 0 ? '⚠ '+t.atRiskCount : '✓ 0'}</span>
            <span class="dl-hdr-stat-lbl">At-Risk</span>
          </div>
        </div>
      </div>
      <table class="dl-table">
        <thead><tr>
          <th>Reporting Period</th>
          <th>Due Date</th>
          <th>Days from Today</th>
          <th>Grants Submitted</th>
          <th>Status</th>
        </tr></thead>
        <tbody>${rows.join('')}</tbody>
      </table>
      <div class="dl-legend">
        <div class="dl-legend-item"><div class="dl-legend-dot" style="background:var(--green)"></div> All submitted</div>
        <div class="dl-legend-item"><div class="dl-legend-dot" style="background:var(--amber)"></div> Partially submitted</div>
        <div class="dl-legend-item"><div class="dl-legend-dot" style="background:var(--red)"></div> Overdue / missing</div>
        <div class="dl-legend-item"><div class="dl-legend-dot" style="background:var(--blue)"></div> Upcoming</div>
      </div>
    </div>`;
  }).join('');

  document.getElementById('v6').innerHTML = `<div class="dl-grid">${cards}</div>`;

  // Animate deadline bars
  setTimeout(() => {
    document.querySelectorAll('#v6 .dl-bar').forEach(el => {
      const w = el.style.width; el.style.width = '0';
      requestAnimationFrame(() => requestAnimationFrame(() => { el.style.width = w; }));
    });
  }, 30);
}

// ─── VIEW 7: LIFECYCLE COMPARISON ─────────────────────────────────────────
function rv7() {
  const rds = Object.entries(DATA.rounds);
  if (!rds.length) {
    document.getElementById('v7').innerHTML = '<div class="empty"><div class="ico">📈</div>No round data.</div>';
    return;
  }

  // Build comparison table rows
  const metrics = [
    { key: 'lifecycle',  label: 'Lifecycle Elapsed',     fmt: v => Math.round(v*100)+'%' },
    { key: 'expected',   label: 'Expected Spend by Now', fmt: v => Math.round(v*100)+'%' },
    { key: 'actual',     label: 'Actual Spend',          fmt: v => Math.round(v*100)+'%' },
    { key: 'pace',       label: 'Pace vs Expected',      fmt: (v,rk,rd) => {
        const delta = v;
        const cls = delta >= -0.05 ? 'br-ok' : delta >= -0.15 ? 'br-warn' : 'br-risk';
        const sign = delta >= 0 ? '+' : '';
        return `<span class="br-badge ${cls}">${sign}${Math.round(delta*100)}pp</span>`;
    }},
    { key: 'budget',     label: 'Total Budget',          fmt: v => '$'+Math.round(v).toLocaleString('en-US') },
    { key: 'spent',      label: 'Total Spent',           fmt: v => '$'+Math.round(v).toLocaleString('en-US') },
    { key: 'unspent',    label: 'Unspent Budget',        fmt: v => '$'+Math.round(v).toLocaleString('en-US') },
    { key: 'grants',     label: 'Total Grants',          fmt: v => v },
    { key: 'atRisk',     label: 'At-Risk Grants',        fmt: (v,rk,rd) => {
        const pct = rd.totals.grantCount > 0 ? Math.round(v/rd.totals.grantCount*100) : 0;
        return v > 0 ? `<span style="color:var(--red);font-weight:700">⚠ ${v} (${pct}%)</span>` : `<span style="color:var(--green);font-weight:700">✓ 0</span>`;
    }},
    { key: 'q1sub',      label: 'Q1 Submitted',          fmt: (v,rk,rd) => `${v}/${rd.totals.grantCount}` },
    { key: 'q2sub',      label: 'Q2 Submitted',          fmt: (v,rk,rd) => `${v}/${rd.totals.grantCount}` },
    { key: 'q3sub',      label: 'Q3 Submitted',          fmt: (v,rk,rd) => `${v}/${rd.totals.grantCount}` },
    { key: 'q4sub',      label: 'Q4 Submitted',          fmt: (v,rk,rd) => `${v}/${rd.totals.grantCount}` },
    { key: 'q5sub',      label: 'Q5 Submitted',          fmt: (v,rk,rd) => `${v}/${rd.totals.grantCount}` },
    { key: 'finalSub',   label: 'Final Report Submitted',fmt: (v,rk,rd) => `${v}/${rd.totals.grantCount}` },
  ];

  // Compute per-round values
  const rdData = rds.map(([rk, rd]) => {
    const start   = new Date(rd.grantFYStart + 'T00:00:00');
    const end     = new Date(rd.finalReportDue + 'T00:00:00');
    const total   = end - start;
    const elapsed = Math.max(0, Math.min(total, TODAY - start));
    const lifecycle = elapsed / total;
    const expected  = lifecycle;
    const actual    = rd.totals.pctSpent;
    const delta     = actual - expected;
    const qc        = rd.totals.quarterlyCounts;
    return { rk, rd,
      lifecycle, expected, actual, pace: delta,
      budget: rd.totals.budget, spent: rd.totals.spent,
      unspent: rd.totals.budget - rd.totals.spent,
      grants: rd.totals.grantCount, atRisk: rd.totals.atRiskCount,
      q1sub: qc[0]||0, q2sub: qc[1]||0, q3sub: qc[2]||0, q4sub: qc[3]||0, q5sub: qc[4]||0,
      finalSub: rd.totals.finalCount,
    };
  });

  const accent = { R6:'#0369A1', R7:'#047857', R8:'#B45309' };

  const hdrCols = rdData.map(({rk, rd}) =>
    `<th style="text-align:center;border-bottom:3px solid ${accent[rk]||'var(--navy)'}">
      <span style="color:${accent[rk]||'var(--navy)'};">${rd.label}</span><br>
      <small style="font-weight:400;color:var(--muted)">Grant FY ${rd.grantFY}</small>
    </th>`
  ).join('');

  const bodyRows = metrics.map(m => {
    const cells = rdData.map(({rk, rd, ...vals}) => {
      const v = vals[m.key];
      const fmt = typeof m.fmt === 'function' ? m.fmt(v, rk, rd) : v;
      return `<td style="text-align:center">${fmt}</td>`;
    }).join('');
    return `<tr><td style="font-weight:600;white-space:nowrap;padding-left:16px">${m.label}</td>${cells}</tr>`;
  }).join('');

  // Lifecycle progress bars per round
  const progressCards = rdData.map(({rk, rd, lifecycle, expected, actual, pace}) => {
    const ac = accent[rk] || 'var(--navy)';
    const paceCls = pace >= -0.05 ? 'br-ok' : pace >= -0.15 ? 'br-warn' : 'br-risk';
    const paceLabel = pace >= -0.05 ? 'On Pace' : pace >= -0.15 ? 'Slightly Behind' : 'Under-Spending';
    const sign = pace >= 0 ? '+' : '';
    return `<div style="background:var(--card);border:2px solid ${ac};border-radius:var(--r);padding:20px;min-width:220px;flex:1">
      <div style="font-size:18px;font-weight:800;color:${ac};margin-bottom:4px">${rd.label}</div>
      <div style="font-size:12px;color:var(--muted);margin-bottom:14px">FY ${rd.grantFY} · ${rd.totals.grantCount} grants</div>
      <div style="margin-bottom:10px">
        <div style="display:flex;justify-content:space-between;font-size:11px;font-weight:700;color:var(--muted);margin-bottom:3px">
          <span>Lifecycle Progress</span><span>${Math.round(lifecycle*100)}%</span>
        </div>
        <div style="background:var(--bg);border-radius:100px;height:8px;overflow:hidden">
          <div style="width:${Math.round(lifecycle*100)}%;height:100%;background:${ac};border-radius:100px"></div>
        </div>
      </div>
      <div style="margin-bottom:10px">
        <div style="display:flex;justify-content:space-between;font-size:11px;font-weight:700;color:var(--muted);margin-bottom:3px">
          <span>Expected Spend</span><span>${Math.round(expected*100)}%</span>
        </div>
        <div style="background:var(--bg);border-radius:100px;height:8px;overflow:hidden">
          <div style="width:${Math.round(expected*100)}%;height:100%;background:#94A3B8;border-radius:100px"></div>
        </div>
      </div>
      <div style="margin-bottom:14px">
        <div style="display:flex;justify-content:space-between;font-size:11px;font-weight:700;color:var(--muted);margin-bottom:3px">
          <span>Actual Spend</span><span>${Math.round(actual*100)}%</span>
        </div>
        <div style="background:var(--bg);border-radius:100px;height:8px;overflow:hidden">
          <div style="width:${Math.min(100,Math.round(actual*100))}%;height:100%;background:${actual < expected - 0.15 ? 'var(--red)' : actual < expected - 0.05 ? 'var(--amber)' : 'var(--green)'};border-radius:100px"></div>
        </div>
      </div>
      <span class="br-badge ${paceCls}">${paceLabel} (${sign}${Math.round(pace*100)}pp)</span>
    </div>`;
  }).join('');

  const html = `
    <div style="margin-bottom:20px">
      <div class="sec-ttl">Round-by-Round Lifecycle Comparison</div>
      <div style="font-size:13px;color:var(--muted);margin-bottom:16px">
        Compares each round's actual spending against where it <em>should</em> be based on time elapsed since the grant award year started.
      </div>
      <div style="display:flex;gap:16px;flex-wrap:wrap;margin-bottom:24px">${progressCards}</div>
    </div>
    <div class="sec-ttl" style="margin-bottom:12px">Detailed Metrics</div>
    <div class="tbl-wrap">
      <table>
        <thead><tr><th style="min-width:200px">Metric</th>${hdrCols}</tr></thead>
        <tbody>${bodyRows}</tbody>
      </table>
    </div>`;

  document.getElementById('v7').innerHTML = html;
}

// ─── RENDER DISPATCHER ────────────────────────────────────────────────────
function render() {
  switch (S.view) {
    case 1: rv1(); break;
    case 2: rv2(); break;
    case 3: rv3(); break;
    case 4: rv4(); break;
    case 5: rv5(); break;
    case 6: rv6(); break;
    case 7: rv7(); break;
  }
}

function animateBars(sel) {
  setTimeout(() => {
    document.querySelectorAll(sel + ' .pb').forEach(el => {
      const w = el.style.width;
      el.style.width = '0';
      requestAnimationFrame(() => requestAnimationFrame(() => { el.style.width = w; }));
    });
  }, 30);
}

// ─── INTERACTION HANDLERS ─────────────────────────────────────────────────
function ds(col) {
  S.sortCol === col ? (S.sortDir *= -1) : (S.sortCol = col, S.sortDir = 1);
  render();
}
function togAlert()      { S.alertOpen = !S.alertOpen; rv4(); }
function setSF(f)        { S.statusFilter = f; rv4(); }
function togCard(pid)    { S.expandedCard = S.expandedCard === pid ? null : pid; rv5(); }
function togRound(rk)    { S.expandedRound = S.expandedRound === rk ? null : rk; rv2(); }

// ─── EVENT LISTENERS ──────────────────────────────────────────────────────
document.querySelectorAll('.tab').forEach(btn => {
  btn.addEventListener('click', () => {
    document.querySelectorAll('.tab').forEach(b => b.classList.remove('on'));
    btn.classList.add('on');
    document.querySelectorAll('.view').forEach(v => v.classList.remove('on'));
    S.view = +btn.dataset.v;
    document.getElementById('v' + S.view).classList.add('on');
    render();
  });
});

document.querySelectorAll('.pill').forEach(p => {
  p.addEventListener('click', () => {
    document.querySelectorAll('.pill').forEach(x => x.classList.remove('on'));
    p.classList.add('on');
    S.round = p.dataset.r;
    S.expandedRound = null;
    render();
  });
});

document.getElementById('srch').addEventListener('input', e => {
  S.search = e.target.value;
  render();
});

// ─── INIT ─────────────────────────────────────────────────────────────────
document.getElementById('gen-ts').textContent = DATA.generated.replace('T', ' ');
renderDeadlineBar();
render();
</script>
</body>
</html>"""


def main():
    if not os.path.exists(SPREADSHEET):
        print(f"ERROR: {SPREADSHEET} not found.")
        return

    today = date.today()
    wb = load_workbook(SPREADSHEET, data_only=True)
    scraped = load_scraped()

    output = {
        'generated': datetime.now().isoformat(timespec='seconds'),
        'rounds': {},
    }

    # Dashboard shows R6, R7, R8 only — R5 is archived (R9 can be added here when ready)
    DASHBOARD_ROUNDS = ['R6', 'R7', 'R8']

    for rk, meta in ROUND_META.items():
        if rk not in DASHBOARD_ROUNDS:
            continue
        if meta['sheet'] not in wb.sheetnames:
            print(f"WARNING: '{meta['sheet']}' not found in workbook, skipping.")
            continue
        ws = wb[meta['sheet']]
        rd = extract_round(ws, rk, meta, today, scraped)
        output['rounds'][rk] = rd
        n = rd['totals']['grantCount']
        at = rd['totals']['atRiskCount']
        print(f"  {meta['label']}: {n} grants, {at} at-risk, "
              f"budget ${rd['totals']['budget']:,.0f}, "
              f"spent ${rd['totals']['spent']:,.0f} ({rd['totals']['pctSpent']*100:.1f}%)")

    # Embed logo as base64 data URI so the HTML is fully self-contained
    logo_path = 'Logo-w-Tagline_CVML_Reversed.svg'
    if os.path.exists(logo_path):
        with open(logo_path, 'rb') as f:
            logo_b64 = base64.b64encode(f.read()).decode('ascii')
        logo_src = f'data:image/svg+xml;base64,{logo_b64}'
    else:
        logo_src = ''

    data_json = json.dumps(output, separators=(',', ':'))
    html = (HTML_TEMPLATE
            .replace('__DATA_JSON__', data_json)
            .replace('__TODAY__', today.isoformat())
            .replace('__LOGO_SRC__', logo_src))

    out_path = 'dashboard.html'
    with open(out_path, 'w', encoding='utf-8') as f:
        f.write(html)

    total = sum(r['totals']['grantCount'] for r in output['rounds'].values())
    print(f"\nGenerated {out_path} with {total} total grants across {len(output['rounds'])} rounds.")
    print(f"Open {out_path} in your browser to view the dashboard.")


if __name__ == '__main__':
    main()
