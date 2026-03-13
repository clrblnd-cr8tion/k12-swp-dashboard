#!/usr/bin/env python3
"""
Regenerate Round 5-8 tabs in the K12 Reporting Status spreadsheet
from scraped_data.json. Preserves Notes and Unexpended columns
by matching on Proposal ID.

R5 uses a 17-column layout (A-Q).
R6-R8 use a 19-column layout (A-S) with Plan Status (L) and Lead LEA (M).

Usage: python3 regenerate_spreadsheet.py
"""
import json
import os
import sys
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from utils import SPREADSHEET, SCRAPED_DATA, is_data_row

# Column layouts (1-based, openpyxl convention)
# Columns 1-11 (A-K) are identical in both layouts.
LAYOUT_17 = {  # R5
    'quarterly_start': 5, 'final_report': 10, 'waiting': 11,
    'plan_status': None, 'lead_lea': None,
    'grant_amount': 12, 'total_exp': 13, 'approved_exp': 14,
    'pct_spent': 15, 'notes': 16, 'unexpended': 17,
    'total_cols': 17,
}

LAYOUT_19 = {  # R6, R7, R8
    'quarterly_start': 5, 'final_report': 10, 'waiting': 11,
    'plan_status': 12, 'lead_lea': 13,
    'grant_amount': 14, 'total_exp': 15, 'approved_exp': 16,
    'pct_spent': 17, 'notes': 18, 'unexpended': 19,
    'total_cols': 19,
}

# Round configurations: sheet name, quarterly column headers, quarterly keys, final report header
ROUND_CONFIGS = {
    'R5': {
        'sheet_name': 'Round 5',
        'layout': LAYOUT_17,
        'quarterly_labels': [
            'FY22-23 Q4\n(08/31/23)', 'FY23-24 Q2\n(02/28/24)',
            'FY23-24 Q4\n(08/31/24)', 'FY24-25 Q2\n(02/28/25)',
            'FY24-25 Q4\n(08/31/25)'
        ],
        'quarterly_keys': [
            'FY22-23_Q4', 'FY23-24_Q2', 'FY23-24_Q4',
            'FY24-25_Q2', 'FY24-25_Q4'
        ],
        'final_report_header': 'Final Report\n(Due 09/30/25)',
    },
    'R6': {
        'sheet_name': 'Round 6',
        'layout': LAYOUT_19,
        'quarterly_labels': [
            'FY23-24 Q4\n(08/31/24)', 'FY24-25 Q2\n(02/28/25)',
            'FY24-25 Q4\n(08/31/25)', 'FY25-26 Q2\n(02/28/26)',
            'FY25-26 Q4\n(08/31/26)'
        ],
        'quarterly_keys': [
            'FY23-24_Q4', 'FY24-25_Q2', 'FY24-25_Q4',
            'FY25-26_Q2', 'FY25-26_Q4'
        ],
        'final_report_header': 'Final Report\n(Due 09/30/26)',
    },
    'R7': {
        'sheet_name': 'Round 7',
        'layout': LAYOUT_19,
        'quarterly_labels': [
            'FY24-25 Q4\n(08/31/25)', 'FY25-26 Q2\n(02/28/26)',
            'FY25-26 Q4\n(08/31/26)', 'FY26-27 Q2\n(02/28/27)',
            'FY26-27 Q4\n(08/31/27)'
        ],
        'quarterly_keys': [
            'FY24-25_Q4', 'FY25-26_Q2', 'FY25-26_Q4',
            'FY26-27_Q2', 'FY26-27_Q4'
        ],
        'final_report_header': 'Final Report\n(Due 09/30/27)',
    },
    'R8': {
        'sheet_name': 'Round 8',
        'layout': LAYOUT_19,
        'quarterly_labels': [
            'FY25-26 Q4\n(08/31/26)', 'FY26-27 Q2\n(02/28/27)',
            'FY26-27 Q4\n(08/31/27)', 'FY27-28 Q2\n(02/28/28)',
            'FY27-28 Q4\n(08/31/28)'
        ],
        'quarterly_keys': [
            'FY25-26_Q4', 'FY26-27_Q2', 'FY26-27_Q4',
            'FY27-28_Q2', 'FY27-28_Q4'
        ],
        'final_report_header': 'Final Report\n(Due 09/30/28)',
    },
}

# Fixed headers for columns A-D and K onward (E-J are per-round)
BASE_HEADERS_17 = [
    'Lead Institution', 'Proposal ID', '# Reporting Insitutions', 'Grant Name',
    None, None, None, None, None,  # E-I: quarterly (filled per round)
    None,  # J: Final Report (filled per round)
    'Report Waiting Approval', 'Grant Amount', 'Total Reported Expenditures',
    'Total Reported Expenditures Approved', '% Spent', 'Notes',
    'Unexpended according to last fiscal report'
]

BASE_HEADERS_19 = [
    'Lead Institution', 'Proposal ID', '# Reporting Insitutions', 'Grant Name',
    None, None, None, None, None,  # E-I: quarterly (filled per round)
    None,  # J: Final Report (filled per round)
    'Report Waiting Approval', 'Plan Status', 'Lead LEA', 'Grant Amount',
    'Total Reported Expenditures', 'Total Reported Expenditures Approved',
    '% Spent', 'Notes', 'Unexpended according to last fiscal report'
]

HEADERS_BY_LAYOUT = {17: BASE_HEADERS_17, 19: BASE_HEADERS_19}

BLUE_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF')
DATA_FONT = Font(name='Arial')
BOLD_FONT = Font(name='Arial', bold=True)

COL_WIDTHS_17 = [28.38, 10.88, 10.88, 59.0, 7.75, 7.75, 7.75, 7.75, 7.75, 7.75,
                 8.75, 11.75, 11.75, 11.75, 8.63, 32.38, 15.0]

COL_WIDTHS_19 = [28.38, 10.88, 10.88, 59.0, 7.75, 7.75, 7.75, 7.75, 7.75, 7.75,
                 8.75, 8.75, 28.38, 11.75, 11.75, 11.75, 8.63, 32.38, 15.0]

WIDTHS_BY_LAYOUT = {17: COL_WIDTHS_17, 19: COL_WIDTHS_19}


def _data_cell(ws, row, column, value, number_format=None):
    """Write a cell with DATA_FONT and optional number format."""
    cell = ws.cell(row=row, column=column, value=value)
    cell.font = DATA_FONT
    if number_format:
        cell.number_format = number_format
    return cell


def regenerate():
    if not os.path.exists(SCRAPED_DATA):
        print(f"ERROR: {SCRAPED_DATA} not found. Run the NOVA scrape first.")
        sys.exit(1)
    if not os.path.exists(SPREADSHEET):
        print(f"ERROR: {SPREADSHEET} not found. Ensure the spreadsheet is in this directory.")
        sys.exit(1)

    with open(SCRAPED_DATA, 'r') as f:
        data = json.load(f)

    wb = load_workbook(SPREADSHEET)

    for round_key in ['R5', 'R6', 'R7', 'R8']:
        if round_key not in data:
            print(f"WARNING: {round_key} not found in {SCRAPED_DATA}, skipping.")
            continue
        config = ROUND_CONFIGS[round_key]
        sheet_name = config['sheet_name']
        round_data = data[round_key]
        grants = round_data['grants']

        L = config['layout']

        # Preserve Notes and Unexpended from existing sheet (find by header text)
        preserved = {}
        if sheet_name in wb.sheetnames:
            old_ws = wb[sheet_name]
            notes_col = unexpended_col = None
            for col_idx in range(1, (old_ws.max_column or 0) + 1):
                hdr = old_ws.cell(row=1, column=col_idx).value
                if hdr == 'Notes':
                    notes_col = col_idx
                if hdr and str(hdr).startswith('Unexpended'):
                    unexpended_col = col_idx
            for row in old_ws.iter_rows(min_row=2, values_only=False):
                pid = row[1].value
                if is_data_row(pid):
                    preserved[str(pid)] = {
                        'notes': row[notes_col - 1].value if notes_col and len(row) >= notes_col else None,
                        'unexpended': row[unexpended_col - 1].value if unexpended_col and len(row) >= unexpended_col else None,
                    }
            idx = wb.sheetnames.index(sheet_name)
            del wb[sheet_name]
            ws = wb.create_sheet(sheet_name, idx)
        else:
            ws = wb.create_sheet(sheet_name)

        # Sort grants alphabetically by lead institution
        grants.sort(key=lambda g: g['leadInstitution'])

        # Set column widths
        col_widths = WIDTHS_BY_LAYOUT[L['total_cols']]
        for i, w in enumerate(col_widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = w

        # Build headers
        headers = list(HEADERS_BY_LAYOUT[L['total_cols']])
        for i, label in enumerate(config['quarterly_labels']):
            headers[4 + i] = label
        headers[9] = config['final_report_header']

        # Write header row
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = BLUE_FILL
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

        # Write data rows
        for row_idx, grant in enumerate(grants, 2):
            pid = grant['planId']
            dc = lambda col, val, fmt=None: _data_cell(ws, row_idx, col, val, fmt)

            dc(1, grant['leadInstitution'])
            dc(2, pid)
            dc(3, grant['dashboardInstitutions'])
            dc(4, grant['grantName'])

            # E-I: Quarterly status
            for qi, qkey in enumerate(config['quarterly_keys']):
                val = grant['quarterlyStatus'].get(qkey, False)
                dc(L['quarterly_start'] + qi, True if val else None)

            # J: Final Report
            final = grant.get('finalReport', False)
            dc(L['final_report'], True if final else None)

            # K: Report Waiting Approval (case-insensitive comparison)
            da = grant.get('dashboardApproval', '')
            da_lower = da.lower()
            waiting = da if da_lower in ('pending approval', 'submitted') else None
            dc(L['waiting'], waiting)

            # Plan Status and Lead LEA (R6-R8 only)
            if L['plan_status']:
                dc(L['plan_status'], grant.get('planStatus', '') or None)
            if L['lead_lea']:
                dc(L['lead_lea'], grant.get('leadLEA', '') or None)

            # Financial columns
            dc(L['grant_amount'], grant['projectBudget'], '#,##0')
            dc(L['total_exp'], grant['ptdExpenditure'], '#,##0')

            approved_exp = grant['ptdExpenditure'] if da_lower in ('approved', 'certified') else 0
            dc(L['approved_exp'], approved_exp, '#,##0')

            # % Spent (formula, guarded against division by zero)
            ga_col = get_column_letter(L['grant_amount'])
            te_col = get_column_letter(L['total_exp'])
            dc(L['pct_spent'],
               f'=IF({ga_col}{row_idx}=0,"",{te_col}{row_idx}/{ga_col}{row_idx})',
               '0.0%')

            # Notes and Unexpended (preserved from previous tab)
            prev = preserved.get(pid, {})
            dc(L['notes'], prev.get('notes'))
            dc(L['unexpended'], prev.get('unexpended'))

        # Total row
        total_row = len(grants) + 2
        ws.cell(row=total_row, column=1, value='Central Mother Lode Region').font = BOLD_FONT

        # COUNTIF for quarterly columns E-I and Final Report J
        for col in range(L['quarterly_start'], L['final_report'] + 1):
            cl = get_column_letter(col)
            ws.cell(row=total_row, column=col,
                    value=f'=COUNTIF({cl}2:{cl}{total_row - 1},TRUE)').font = BOLD_FONT

        # SUM for Grant Amount, Total Exp, Approved Exp
        for col in [L['grant_amount'], L['total_exp'], L['approved_exp']]:
            cl = get_column_letter(col)
            cell = ws.cell(row=total_row, column=col,
                           value=f'=SUM({cl}2:{cl}{total_row - 1})')
            cell.font = BOLD_FONT
            cell.number_format = '#,##0'

        # % Spent for total row (guarded against division by zero)
        ga_col = get_column_letter(L['grant_amount'])
        te_col = get_column_letter(L['total_exp'])
        cell = ws.cell(row=total_row, column=L['pct_spent'],
                       value=f'=IF({ga_col}{total_row}=0,"",{te_col}{total_row}/{ga_col}{total_row})')
        cell.font = BOLD_FONT
        cell.number_format = '0.0%'

        # Report
        restored = sum(1 for g in grants if g['planId'] in preserved)
        print(f"{sheet_name}: {len(grants)} grants, total row at {total_row}, "
              f"preserved {restored}/{len(preserved)} Notes/Unexpended")

    wb.save(SPREADSHEET)
    print(f"\nSaved {SPREADSHEET}")


if __name__ == '__main__':
    regenerate()
