#!/usr/bin/env python3
"""
Automated validation checks for K12 SWP grant tracking data.

Validates scraped_data.json and the spreadsheet against the validation
checklist in CLAUDE.md. Prints PASS/FAIL/WARNING for each check.

Usage: python3 validate.py
"""
import json
import sys
import os
from datetime import date
from openpyxl import load_workbook

from utils import SPREADSHEET, SCRAPED_DATA, is_data_row

# Quarterly deadlines per round (Col E through Col I)
QUARTERLY_DEADLINES = {
    'R5': [
        ('FY22-23_Q4', date(2023, 8, 31)),
        ('FY23-24_Q2', date(2024, 2, 28)),
        ('FY23-24_Q4', date(2024, 8, 31)),
        ('FY24-25_Q2', date(2025, 2, 28)),
        ('FY24-25_Q4', date(2025, 8, 31)),
    ],
    'R6': [
        ('FY23-24_Q4', date(2024, 8, 31)),
        ('FY24-25_Q2', date(2025, 2, 28)),
        ('FY24-25_Q4', date(2025, 8, 31)),
        ('FY25-26_Q2', date(2026, 2, 28)),
        ('FY25-26_Q4', date(2026, 8, 31)),
    ],
    'R7': [
        ('FY24-25_Q4', date(2025, 8, 31)),
        ('FY25-26_Q2', date(2026, 2, 28)),
        ('FY25-26_Q4', date(2026, 8, 31)),
        ('FY26-27_Q2', date(2027, 2, 28)),
        ('FY26-27_Q4', date(2027, 8, 31)),
    ],
    'R8': [
        ('FY25-26_Q4', date(2026, 8, 31)),
        ('FY26-27_Q2', date(2027, 2, 28)),
        ('FY26-27_Q4', date(2027, 8, 31)),
        ('FY27-28_Q2', date(2028, 2, 28)),
        ('FY27-28_Q4', date(2028, 8, 31)),
    ],
}

FINAL_REPORT_DEADLINES = {
    'R5': date(2025, 9, 30),
    'R6': date(2026, 9, 30),
    'R7': date(2027, 9, 30),
    'R8': date(2028, 9, 30),
}

ROUND_SHEET_NAMES = {
    'R5': 'Round 5',
    'R6': 'Round 6',
    'R7': 'Round 7',
    'R8': 'Round 8',
}

passed = 0
failed = 0
warnings = 0


def result(status, check_num, description, detail=""):
    global passed, failed, warnings
    if status == "PASS":
        passed += 1
        icon = "PASS"
    elif status == "FAIL":
        failed += 1
        icon = "FAIL"
    else:
        warnings += 1
        icon = "WARN"
    msg = f"  [{icon}] #{check_num}: {description}"
    if detail:
        msg += f"\n         {detail}"
    print(msg)


def load_data():
    if not os.path.exists(SCRAPED_DATA):
        print(f"ERROR: {SCRAPED_DATA} not found. Run the NOVA scrape first.")
        sys.exit(1)
    with open(SCRAPED_DATA, 'r') as f:
        return json.load(f)


def load_spreadsheet():
    if not os.path.exists(SPREADSHEET):
        print(f"ERROR: {SPREADSHEET} not found.")
        sys.exit(1)
    return load_workbook(SPREADSHEET, data_only=True)


def check_3_budget_sanity(data):
    """Check 3: All projectBudget values > $0"""
    issues = []
    for rk, rd in data.items():
        for g in rd['grants']:
            if g['projectBudget'] <= 0:
                issues.append(f"{rk} {g['planId']} ({g['grantName']}): ${g['projectBudget']}")
    if issues:
        result("WARN", 3, "Budget sanity — some grants have $0 budget", "; ".join(issues))
    else:
        result("PASS", 3, "Budget sanity — all projectBudget > $0")


def check_4_expenditure_bounds(data):
    """Check 4: 0 <= ptdExpenditure <= projectBudget"""
    issues = []
    for rk, rd in data.items():
        for g in rd['grants']:
            if g['ptdExpenditure'] < 0:
                issues.append(f"{rk} {g['planId']}: negative expenditure ${g['ptdExpenditure']}")
            elif g['ptdExpenditure'] > g['projectBudget'] and g['projectBudget'] > 0:
                pct = g['ptdExpenditure'] / g['projectBudget'] * 100
                issues.append(f"{rk} {g['planId']}: {pct:.1f}% spent (over budget)")
    if issues:
        result("WARN", 4, "Expenditure bounds — some grants over budget", "; ".join(issues))
    else:
        result("PASS", 4, "Expenditure bounds — all within range")


def check_5_budget_remaining(data):
    """Check 5: budgetRemaining ~= projectBudget - ptdExpenditure"""
    issues = []
    for rk, rd in data.items():
        for g in rd['grants']:
            expected = g['projectBudget'] - g['ptdExpenditure']
            actual = g['budgetRemaining']
            if abs(expected - actual) > 1:
                issues.append(f"{rk} {g['planId']}: remaining=${actual}, expected=${expected}")
    if issues:
        result("FAIL", 5, "Budget remaining mismatch", "; ".join(issues[:5]))
    else:
        result("PASS", 5, "Budget remaining — all values consistent")


def check_6_spreadsheet_row_count(data, wb):
    """Check 6: Data rows per tab must match expected grant count"""
    issues = []
    for rk, rd in data.items():
        sheet_name = ROUND_SHEET_NAMES.get(rk)
        if not sheet_name or sheet_name not in wb.sheetnames:
            issues.append(f"{rk}: sheet '{sheet_name}' not found")
            continue
        ws = wb[sheet_name]
        # Count data rows (skip header row 1, skip total row at bottom)
        data_rows = 0
        for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
            pid = row[1]
            if is_data_row(pid):
                data_rows += 1
        expected = len(rd['grants'])
        if data_rows != expected:
            issues.append(f"{rk}: {data_rows} rows in sheet vs {expected} grants in JSON")
    if issues:
        result("FAIL", 6, "Spreadsheet row count mismatch", "; ".join(issues))
    else:
        result("PASS", 6, "Spreadsheet row count — matches grant counts")


def check_7_column_headers(wb):
    """Check 7: Column headers match expected layout from ROUND_CONFIGS"""
    from regenerate_spreadsheet import ROUND_CONFIGS, HEADERS_BY_LAYOUT
    issues = []
    for rk, config in ROUND_CONFIGS.items():
        sheet_name = config['sheet_name']
        if sheet_name not in wb.sheetnames:
            issues.append(f"{rk}: sheet '{sheet_name}' not found")
            continue
        ws = wb[sheet_name]
        L = config['layout']
        # Build expected headers using the correct layout
        expected = list(HEADERS_BY_LAYOUT[L['total_cols']])
        for i, label in enumerate(config['quarterly_labels']):
            expected[4 + i] = label
        expected[9] = config['final_report_header']
        # Compare
        for col_idx, exp in enumerate(expected):
            actual = ws.cell(row=1, column=col_idx + 1).value
            if actual != exp:
                col_letter = chr(65 + col_idx) if col_idx < 26 else f"Col{col_idx+1}"
                issues.append(f"{rk} {col_letter}: expected '{exp}', got '{actual}'")
    if issues:
        result("FAIL", 7, "Column header mismatch", "; ".join(issues[:10]))
    else:
        result("PASS", 7, "Column headers — all match expected layout")


def check_8_sum_validation(data):
    """Check 8: Print total budget and expenditure per round"""
    print("\n  --- Sum Validation (Check #8) ---")
    for rk in ['R5', 'R6', 'R7', 'R8']:
        if rk not in data:
            continue
        rd = data[rk]
        total_budget = sum(g['projectBudget'] for g in rd['grants'])
        total_exp = sum(g['ptdExpenditure'] for g in rd['grants'])
        pct = (total_exp / total_budget * 100) if total_budget > 0 else 0
        print(f"  {rk}: {len(rd['grants'])} grants, "
              f"Budget=${total_budget:,.0f}, "
              f"Expenditure=${total_exp:,.0f} ({pct:.1f}%)")
    result("PASS", 8, "Sum validation — totals printed above for review")


def check_13_quarterly_monotonicity(data):
    """Check 13: Cross-FY monotonicity (within-FY Q2/Q4 are independent)"""
    issues = []
    for rk, rd in data.items():
        deadlines = QUARTERLY_DEADLINES.get(rk, [])
        if len(deadlines) < 4:
            continue
        for g in rd['grants']:
            qs = g.get('quarterlyStatus', {})
            keys = [d[0] for d in deadlines]
            vals = [qs.get(k, False) for k in keys]
            # Group by FY: keys are like FY24-25_Q4, FY25-26_Q2, FY25-26_Q4, ...
            # Cross-FY check: if any quarter in a later FY is TRUE,
            # all quarters in earlier FYs should be TRUE
            fys = {}
            for k, v in zip(keys, vals):
                fy = k.rsplit('_', 1)[0]  # e.g., "FY24-25"
                fys.setdefault(fy, []).append((k, v))

            fy_list = list(fys.keys())
            for i in range(1, len(fy_list)):
                later_fy = fy_list[i]
                later_has_true = any(v for _, v in fys[later_fy])
                if later_has_true:
                    for j in range(i):
                        earlier_fy = fy_list[j]
                        for k, v in fys[earlier_fy]:
                            if not v:
                                issues.append(
                                    f"{rk} {g['planId']}: {k}=FALSE but {later_fy} has TRUE"
                                )
    if issues:
        detail = "; ".join(issues[:10])
        if len(issues) > 10:
            detail += f" ...and {len(issues) - 10} more"
        result("WARN", 13, "Cross-FY monotonicity violations found", detail)
    else:
        result("PASS", 13, "Cross-FY quarterly monotonicity — consistent")


def check_15_final_report_consistency(data):
    """Check 15: If all quarterly cols TRUE and final report due date passed, Col J should be TRUE"""
    today = date.today()
    issues = []
    for rk, rd in data.items():
        due = FINAL_REPORT_DEADLINES.get(rk)
        if not due or today < due:
            continue
        deadlines = QUARTERLY_DEADLINES.get(rk, [])
        keys = [d[0] for d in deadlines]
        for g in rd['grants']:
            qs = g.get('quarterlyStatus', {})
            all_true = all(qs.get(k, False) for k in keys)
            if all_true and not g.get('finalReport', False):
                issues.append(f"{rk} {g['planId']}: all quarters TRUE but finalReport=FALSE")
    if issues:
        result("WARN", 15, "Final report status inconsistency", "; ".join(issues[:10]))
    else:
        result("PASS", 15, "Final report status — consistent")


def check_16_cross_round_dedup(data):
    """Check 16: No plan ID appears in multiple rounds"""
    seen = {}
    dupes = []
    for rk, rd in data.items():
        for g in rd['grants']:
            pid = g['planId']
            if pid in seen:
                dupes.append(f"{pid} in both {seen[pid]} and {rk}")
            else:
                seen[pid] = rk
    if dupes:
        result("FAIL", 16, "Cross-round duplicate plan IDs found", "; ".join(dupes))
    else:
        result("PASS", 16, "Cross-round deduplication — no duplicates")


def check_18_stale_data(data):
    """Check 18: Flag grants with past-due quarterly deadlines where column is FALSE"""
    today = date.today()
    stale = []
    for rk, rd in data.items():
        deadlines = QUARTERLY_DEADLINES.get(rk, [])
        for g in rd['grants']:
            qs = g.get('quarterlyStatus', {})
            for key, deadline in deadlines:
                if today > deadline and not qs.get(key, False):
                    stale.append(f"{rk} {g['planId']} {key}")
    if stale:
        detail = f"{len(stale)} stale entries. First 10: " + "; ".join(stale[:10])
        result("WARN", 18, "Stale data detected — past-due quarters still FALSE", detail)
    else:
        result("PASS", 18, "No stale data detected")


def main():
    print("K12 SWP Grant Tracking — Validation Report")
    print("=" * 50)

    data = load_data()
    wb = load_spreadsheet()

    print(f"\nRounds in scraped data: {', '.join(data.keys())}")
    print()

    # Automated checks
    check_3_budget_sanity(data)
    check_4_expenditure_bounds(data)
    check_5_budget_remaining(data)
    check_6_spreadsheet_row_count(data, wb)
    check_7_column_headers(wb)
    check_8_sum_validation(data)
    check_13_quarterly_monotonicity(data)
    check_15_final_report_consistency(data)
    check_16_cross_round_dedup(data)
    check_18_stale_data(data)

    # Manual checks reminder
    print("\n  --- Manual Checks (perform these yourself) ---")
    print("  [ ] #9:  Spot check — re-visit 3 random grant pages and compare fields")
    print("  [ ] #10: Formula check — verify % Spent formula for 5 random rows in Excel")
    print("  [ ] #11: Data preservation — verify Notes and Unexpended columns restored after tab rebuild")
    print("  [ ] #12: Quarterly status consistency — compare approvalStatus vs Col E for flagged grants")

    # Summary
    print(f"\n{'=' * 50}")
    print(f"Results: {passed} passed, {failed} failed, {warnings} warnings")
    if failed > 0:
        print("ACTION REQUIRED: Fix failed checks before saving.")
        sys.exit(1)
    elif warnings > 0:
        print("Review warnings above. They may be expected for early-lifecycle grants.")
    else:
        print("All automated checks passed.")


if __name__ == '__main__':
    main()
